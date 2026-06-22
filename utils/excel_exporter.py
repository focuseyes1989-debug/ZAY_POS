# utils/excel_exporter.py
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelExporter:
    """Utility class for exporting data to Excel (.xlsx) format"""
    
    # Define styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=14, color="2c3e50")
    SUBTITLE_FONT = Font(size=10, color="7f8c8d")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @classmethod
    def save_file_dialog(cls, parent, default_name, title="Export to Excel"):
        """Show save file dialog and return file path"""
        file_path, _ = QFileDialog.getSaveFileName(
            parent, title, default_name, "Excel Files (*.xlsx)"
        )
        return file_path
    
    @classmethod
    def create_workbook(cls, title, sheet_name="Report"):
        """Create a new workbook with styled title"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = cls.TITLE_FONT
        title_cell.alignment = cls.CENTER_ALIGN
        
        # Add generation date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = cls.SUBTITLE_FONT
        
        return wb, ws
    
    @classmethod
    def apply_header_style(cls, ws, row, columns):
        """Apply header style to a row"""
        for col, value in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.CENTER_ALIGN
            cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_cell_style(cls, ws, row, col, value, alignment="left"):
        """Apply style to a data cell"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = cls.THIN_BORDER
        
        if alignment == "center":
            cell.alignment = cls.CENTER_ALIGN
        elif alignment == "right":
            cell.alignment = cls.RIGHT_ALIGN
        else:
            cell.alignment = cls.LEFT_ALIGN
        
        return cell
    
    @classmethod
    def auto_adjust_columns(cls, ws, data_rows, start_row=4, start_col=1):
        """Auto adjust column widths based on content"""
        for col in range(start_col, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            # Check header
            header_cell = ws.cell(row=start_row - 1, column=col)
            if header_cell.value:
                max_length = len(str(header_cell.value))
            
            # Check data rows
            for row in range(start_row, start_row + len(data_rows)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @classmethod
    def add_summary_row(cls, ws, row, labels_values):
        """Add a summary row with bold labels"""
        for col, (label, value) in enumerate(labels_values, 1):
            label_cell = ws.cell(row=row, column=col, value=label)
            label_cell.font = Font(bold=True)
            label_cell.border = cls.THIN_BORDER
            
            value_cell = ws.cell(row=row + 1, column=col, value=value)
            value_cell.font = Font(bold=True)
            value_cell.border = cls.THIN_BORDER
    
    @classmethod
    def show_success_message(cls, parent, file_path):
        """Show success message after export"""
        lang = cls._get_lang()
        if lang == "my":
            msg = f"Excel ဖိုင် အောင်မြင်စွာ သိမ်းဆည်းပြီးပါပြီ:\n{file_path}"
        else:
            msg = f"Excel file exported successfully to:\n{file_path}"
        QMessageBox.information(parent, "Export Complete", msg)
    
    @classmethod
    def show_error_message(cls, parent, error):
        """Show error message"""
        lang = cls._get_lang()
        if lang == "my":
            msg = f"Excel ဖိုင် ထုတ်ယူရာတွင် အမှားရှိပါသည်: {error}"
        else:
            msg = f"Failed to export Excel file: {error}"
        QMessageBox.critical(parent, "Export Error", msg)
    
    @classmethod
    def _get_lang(cls):
        try:
            from models.database import connect_db
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"