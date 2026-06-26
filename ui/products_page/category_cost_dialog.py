# ui/products_page/category_cost_dialog.py
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QLabel, QMessageBox, QFrame, QLineEdit, QComboBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.language import lang
from utils.excel_exporter import ExcelExporter
from ui.widgets.pagination_widget import PaginationWidget
from datetime import datetime
from loguru import logger


class CategoryCostDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Category Cost Breakdown")
        self.setMinimumSize(900, 550)
        self.setModal(True)
        
        # Store data for export
        self.category_data = []
        self.filtered_data = []
        self.total_cost_all = 0
        
        # Pagination variables
        self.current_page = 1
        self.page_size = 50
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # Top bar with export button
        top_layout = QHBoxLayout()
        
        # Export button on the right
        self.btn_export = QPushButton("Export Excel")
        self.btn_export.clicked.connect(self.export_to_excel)
        top_layout.addStretch()
        top_layout.addWidget(self.btn_export)
        
        layout.addLayout(top_layout)
        
        # Summary Cards - Dashboard card style with proper theme support
        card_layout = QHBoxLayout()
        card_layout.setSpacing(15)
        
        # Card 1: Total Categories
        self.card_categories = QFrame()
        self.card_categories.setObjectName("dashboardCard")
        self.card_categories.setFixedHeight(80)
        self.card_categories.setStyleSheet("""
            QFrame#dashboardCard {
                background-color: palette(base);
                border: 1px solid palette(mid);
                border-radius: 8px;
                padding: 10px;
            }
            QFrame#dashboardCard:hover {
                border: 1px solid palette(highlight);
            }
        """)
        card1_layout = QVBoxLayout()
        card1_layout.setContentsMargins(10, 8, 10, 8)
        card1_layout.setSpacing(2)
        self.card_categories_label = QLabel("Total Categories")
        self.card_categories_label.setObjectName("cardTitle")
        self.card_categories_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_categories_label.setStyleSheet("color: palette(windowText); font-size: 10pt;")
        self.card_categories_value = QLabel("0")
        self.card_categories_value.setObjectName("cardValue")
        self.card_categories_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_categories_value.setStyleSheet("color: palette(windowText); font-size: 18pt; font-weight: bold;")
        card1_layout.addWidget(self.card_categories_label)
        card1_layout.addWidget(self.card_categories_value)
        self.card_categories.setLayout(card1_layout)
        card_layout.addWidget(self.card_categories, 1)
        
        # Card 2: Total Products
        self.card_products = QFrame()
        self.card_products.setObjectName("dashboardCard")
        self.card_products.setFixedHeight(80)
        self.card_products.setStyleSheet("""
            QFrame#dashboardCard {
                background-color: palette(base);
                border: 1px solid palette(mid);
                border-radius: 8px;
                padding: 10px;
            }
            QFrame#dashboardCard:hover {
                border: 1px solid palette(highlight);
            }
        """)
        card2_layout = QVBoxLayout()
        card2_layout.setContentsMargins(10, 8, 10, 8)
        card2_layout.setSpacing(2)
        self.card_products_label = QLabel("Total Products")
        self.card_products_label.setObjectName("cardTitle")
        self.card_products_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_products_label.setStyleSheet("color: palette(windowText); font-size: 10pt;")
        self.card_products_value = QLabel("0")
        self.card_products_value.setObjectName("cardValue")
        self.card_products_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_products_value.setStyleSheet("color: palette(windowText); font-size: 18pt; font-weight: bold;")
        card2_layout.addWidget(self.card_products_label)
        card2_layout.addWidget(self.card_products_value)
        self.card_products.setLayout(card2_layout)
        card_layout.addWidget(self.card_products, 1)
        
        # Card 3: Total Cost
        self.card_cost = QFrame()
        self.card_cost.setObjectName("dashboardCard")
        self.card_cost.setFixedHeight(80)
        self.card_cost.setStyleSheet("""
            QFrame#dashboardCard {
                background-color: palette(base);
                border: 1px solid palette(mid);
                border-radius: 8px;
                padding: 10px;
            }
            QFrame#dashboardCard:hover {
                border: 1px solid palette(highlight);
            }
        """)
        card3_layout = QVBoxLayout()
        card3_layout.setContentsMargins(10, 8, 10, 8)
        card3_layout.setSpacing(2)
        self.card_cost_label = QLabel("Total Cost")
        self.card_cost_label.setObjectName("cardTitle")
        self.card_cost_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_cost_label.setStyleSheet("color: palette(windowText); font-size: 10pt;")
        self.card_cost_value = QLabel("0")
        self.card_cost_value.setObjectName("cardValue")
        self.card_cost_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_cost_value.setStyleSheet("color: palette(windowText); font-size: 18pt; font-weight: bold;")
        card3_layout.addWidget(self.card_cost_label)
        card3_layout.addWidget(self.card_cost_value)
        self.card_cost.setLayout(card3_layout)
        card_layout.addWidget(self.card_cost, 1)
        
        # Card 4: Total Stock
        self.card_stock = QFrame()
        self.card_stock.setObjectName("dashboardCard")
        self.card_stock.setFixedHeight(80)
        self.card_stock.setStyleSheet("""
            QFrame#dashboardCard {
                background-color: palette(base);
                border: 1px solid palette(mid);
                border-radius: 8px;
                padding: 10px;
            }
            QFrame#dashboardCard:hover {
                border: 1px solid palette(highlight);
            }
        """)
        card4_layout = QVBoxLayout()
        card4_layout.setContentsMargins(10, 8, 10, 8)
        card4_layout.setSpacing(2)
        self.card_stock_label = QLabel("Total Stock")
        self.card_stock_label.setObjectName("cardTitle")
        self.card_stock_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_stock_label.setStyleSheet("color: palette(windowText); font-size: 10pt;")
        self.card_stock_value = QLabel("0")
        self.card_stock_value.setObjectName("cardValue")
        self.card_stock_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.card_stock_value.setStyleSheet("color: palette(windowText); font-size: 18pt; font-weight: bold;")
        card4_layout.addWidget(self.card_stock_label)
        card4_layout.addWidget(self.card_stock_value)
        self.card_stock.setLayout(card4_layout)
        card_layout.addWidget(self.card_stock, 1)
        
        layout.addLayout(card_layout)
        
        # Filter section
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search category...")
        self.search_input.textChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_input, 2)
        
        # Min percentage filter
        self.min_percentage_filter = QComboBox()
        self.min_percentage_filter.addItems(["All", "> 10%", "> 25%", "> 50%"])
        self.min_percentage_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(QLabel("Min %:"))
        filter_layout.addWidget(self.min_percentage_filter, 1)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Category", "Product Count", "Total Cost", 
            "Total Stock", "Stock Value (Cost)", "Percentage"
        ])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        
        # Set column stretch modes
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        
        # Set uniform row height
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.verticalHeader().setMinimumSectionSize(25)
        self.table.verticalHeader().setMaximumSectionSize(40)
        self.table.setWordWrap(False)
        
        layout.addWidget(self.table)
        
        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        self.setLayout(layout)
        
        # Load data
        self.load_data()
        
        # Language support
        lang.language_changed.connect(self.retranslateUi)
        self.retranslateUi()
        
        # Set window flags
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint
        )
    
    def retranslateUi(self):
        """Update UI text when language changes"""
        lang_code = lang.get_current()
        if lang_code == "my":
            self.setWindowTitle("အမျိုးအစားအလိုက် အရင်းငွေစာရင်း")
            self.btn_export.setText("Excel ထုတ်မည်")
            self.search_input.setPlaceholderText("အမျိုးအစားရှာရန်...")
            self.min_percentage_filter.setItemText(0, "အားလုံး")
            self.min_percentage_filter.setItemText(1, "> 10%")
            self.min_percentage_filter.setItemText(2, "> 25%")
            self.min_percentage_filter.setItemText(3, "> 50%")
            self.card_categories_label.setText("စုစုပေါင်း အမျိုးအစား")
            self.card_products_label.setText("စုစုပေါင်း ပစ္စည်း")
            self.card_cost_label.setText("စုစုပေါင်း ကုန်ကျငွေ")
            self.card_stock_label.setText("စုစုပေါင်း စတော့")
            self.table.setHorizontalHeaderLabels([
                "အမျိုးအစား", "ပစ္စည်းအရေအတွက်", "စုစုပေါင်းကုန်ကျငွေ", 
                "စုစုပေါင်းစတော့", "စတော့တန်ဖိုး", "ရာခိုင်နှုန်း"
            ])
        else:
            self.setWindowTitle("Category Cost Breakdown")
            self.btn_export.setText("Export Excel")
            self.search_input.setPlaceholderText("Search category...")
            self.min_percentage_filter.setItemText(0, "All")
            self.min_percentage_filter.setItemText(1, "> 10%")
            self.min_percentage_filter.setItemText(2, "> 25%")
            self.min_percentage_filter.setItemText(3, "> 50%")
            self.card_categories_label.setText("Total Categories")
            self.card_products_label.setText("Total Products")
            self.card_cost_label.setText("Total Cost")
            self.card_stock_label.setText("Total Stock")
            self.table.setHorizontalHeaderLabels([
                "Category", "Product Count", "Total Cost", 
                "Total Stock", "Stock Value (Cost)", "Percentage"
            ])
        
        # Update card values
        self.update_summary_cards()
    
    def on_filter_changed(self):
        """Handle filter changes"""
        self.current_page = 1
        self.apply_filters_and_pagination()
    
    def on_page_changed(self, page: int, page_size: int):
        """Handle page changes from pagination widget"""
        self.current_page = page
        self.page_size = page_size
        self.apply_filters_and_pagination()
    
    def load_data(self):
        """Load category cost data from database"""
        conn = connect_db()
        cursor = conn.cursor()
        
        # Get total cost for all products (for percentage calculation)
        cursor.execute("""
            SELECT SUM(COALESCE(cost, 0) * COALESCE(stock, 0))
            FROM products
            WHERE (sold_by IS NULL OR sold_by != 'Service')
              AND COALESCE(stock, 0) > 0
        """)
        self.total_cost_all = cursor.fetchone()[0] or 0
        
        # Get category breakdown
        cursor.execute("""
            SELECT 
                COALESCE(category, 'Uncategorized') as category,
                COUNT(*) as product_count,
                SUM(COALESCE(cost, 0) * COALESCE(stock, 0)) as total_cost,
                SUM(COALESCE(stock, 0)) as total_stock,
                SUM(COALESCE(cost, 0) * COALESCE(stock, 0)) as stock_value
            FROM products
            WHERE (sold_by IS NULL OR sold_by != 'Service')
              AND COALESCE(stock, 0) > 0
            GROUP BY category
            ORDER BY total_cost DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        
        self.category_data = rows
        self.apply_filters_and_pagination()
        self.update_summary_cards()
    
    def apply_filters_and_pagination(self):
        """Apply search filter, percentage filter and pagination"""
        search_text = self.search_input.text().strip().lower()
        min_percentage = self.min_percentage_filter.currentText()
        
        # Parse min percentage value
        min_percent_value = 0
        if min_percentage == "> 10%":
            min_percent_value = 10
        elif min_percentage == "> 25%":
            min_percent_value = 25
        elif min_percentage == "> 50%":
            min_percent_value = 50
        
        # Filter rows
        filtered_rows = []
        for row in self.category_data:
            category, product_count, total_cost, total_stock, stock_value = row
            
            # Search filter
            if search_text:
                if search_text not in category.lower():
                    continue
            
            # Percentage filter
            if min_percent_value > 0 and self.total_cost_all > 0:
                percentage = (total_cost / self.total_cost_all) * 100
                if percentage <= min_percent_value:
                    continue
            
            filtered_rows.append(row)
        
        self.filtered_data = filtered_rows
        
        # Update total items
        total_items = len(filtered_rows)
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        # Calculate pagination
        start_idx = (self.current_page - 1) * self.page_size
        end_idx = min(start_idx + self.page_size, total_items)
        page_rows = filtered_rows[start_idx:end_idx]
        
        # Populate table
        self.populate_table(page_rows)
    
    def populate_table(self, rows):
        """Populate table with category data"""
        self.table.setRowCount(len(rows))
        symbol = get_currency_symbol()
        
        grand_total = self.total_cost_all
        
        for row_idx, row in enumerate(rows):
            category, product_count, total_cost, total_stock, stock_value = row
            
            # Category
            category_item = QTableWidgetItem(category)
            category_item.setFlags(category_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 0, category_item)
            
            # Product Count
            count_item = QTableWidgetItem(str(product_count))
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            count_item.setFlags(count_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 1, count_item)
            
            # Total Cost
            cost_item = QTableWidgetItem(format_money(total_cost, symbol))
            cost_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            cost_item.setFlags(cost_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 2, cost_item)
            
            # Total Stock
            stock_item = QTableWidgetItem(str(total_stock))
            stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            stock_item.setFlags(stock_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 3, stock_item)
            
            # Stock Value (Cost)
            value_item = QTableWidgetItem(format_money(stock_value, symbol))
            value_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            value_item.setFlags(value_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 4, value_item)
            
            # Percentage
            if grand_total > 0:
                percentage = (total_cost / grand_total) * 100
                percentage_text = f"{percentage:.1f}%"
            else:
                percentage_text = "0%"
            
            percentage_item = QTableWidgetItem(percentage_text)
            percentage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            percentage_item.setFlags(percentage_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            
            # Color coding
            if percentage > 50:
                percentage_item.setForeground(QColor(200, 50, 50))
            elif percentage > 25:
                percentage_item.setForeground(QColor(200, 140, 30))
            elif percentage > 10:
                percentage_item.setForeground(QColor(40, 100, 200))
            else:
                percentage_item.setForeground(QColor(40, 160, 60))
            
            self.table.setItem(row_idx, 5, percentage_item)
        
        # Set uniform row height for all rows
        for row in range(self.table.rowCount()):
            self.table.setRowHeight(row, 30)
    
    def update_summary_cards(self):
        """Update summary card values"""
        symbol = get_currency_symbol()
        
        total_categories = len(self.category_data)
        total_products = sum(row[1] for row in self.category_data)
        total_cost = sum(row[2] for row in self.category_data)
        total_stock = sum(row[3] for row in self.category_data)
        
        self.card_categories_value.setText(str(total_categories))
        self.card_products_value.setText(str(total_products))
        self.card_cost_value.setText(format_money(total_cost, symbol))
        self.card_stock_value.setText(str(total_stock))
    
    def export_to_excel(self):
        """Export category cost breakdown to Excel file"""
        if not self.category_data:
            QMessageBox.warning(self, "No Data", "There is no data to export.")
            return
        
        # Use filtered data if available, otherwise use all data
        export_data = self.filtered_data if self.filtered_data else self.category_data
        
        if not export_data:
            QMessageBox.warning(self, "No Data", "There is no data to export.")
            return
        
        lang_code = lang.get_current()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"category_cost_breakdown_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Category Cost Breakdown" if lang_code != "my" else "အမျိုးအစားအလိုက် အရင်းငွေစာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Category Cost Breakdown"
            
            # Title
            ws.merge_cells('A1:F1')
            if lang_code == "my":
                title_text = "အမျိုးအစားအလိုက် အရင်းငွေစာရင်း"
            else:
                title_text = "CATEGORY COST BREAKDOWN"
            ws['A1'] = title_text
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Subtitle with filter info
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            
            # Filter info
            filter_info = []
            search_text = self.search_input.text().strip()
            if search_text:
                filter_info.append(f"Search: {search_text}")
            min_percentage = self.min_percentage_filter.currentText()
            if min_percentage != "All" and min_percentage != "အားလုံး":
                filter_info.append(f"Min %: {min_percentage}")
            
            if filter_info:
                ws['A3'] = " | ".join(filter_info)
                ws['A3'].font = Font(size=9, italic=True, color="7f8c8d")
                start_row = 5
            else:
                start_row = 4
            
            # Headers
            if lang_code == "my":
                headers = ["အမျိုးအစား", "ပစ္စည်းအရေအတွက်", "စုစုပေါင်းကုန်ကျငွေ", 
                          "စုစုပေါင်းစတော့", "စတော့တန်ဖိုး", "ရာခိုင်နှုန်း"]
            else:
                headers = ["Category", "Product Count", "Total Cost", 
                          "Total Stock", "Stock Value (Cost)", "Percentage"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row_idx, row_data in enumerate(export_data, start=start_row + 1):
                category, product_count, total_cost, total_stock, stock_value = row_data
                
                ws.cell(row=row_idx, column=1, value=category)
                ws.cell(row=row_idx, column=2, value=product_count)
                ws.cell(row=row_idx, column=3, value=format_money(total_cost, symbol))
                ws.cell(row=row_idx, column=4, value=total_stock)
                ws.cell(row=row_idx, column=5, value=format_money(stock_value, symbol))
                
                if self.total_cost_all > 0:
                    percentage = (total_cost / self.total_cost_all) * 100
                    ws.cell(row=row_idx, column=6, value=f"{percentage:.1f}%")
                else:
                    ws.cell(row=row_idx, column=6, value="0%")
            
            # Summary section
            summary_row = len(export_data) + start_row + 2
            total_products = sum(row[1] for row in export_data)
            total_cost = sum(row[2] for row in export_data)
            total_stock = sum(row[3] for row in export_data)
            
            ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
            
            summary_data = [
                ("Total Categories", len(export_data)),
                ("Total Products", total_products),
                ("Total Cost", format_money(total_cost, symbol)),
                ("Total Stock", total_stock)
            ]
            
            for i, (label, value) in enumerate(summary_data):
                row = summary_row + 2 + i
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            
            # Auto adjust columns
            for col in range(1, 7):
                column_letter = chr(64 + col)
                ws.column_dimensions[column_letter].width = 20
            ws.column_dimensions['A'].width = 30
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            logger.error(f"Export category cost breakdown failed: {e}")
            ExcelExporter.show_error_message(self, e)
    
    def keyPressEvent(self, event):
        """Handle keyboard shortcuts"""
        if event.key() == Qt.Key.Key_Escape:
            self.accept()
        elif event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.key() == Qt.Key.Key_E:
                self.export_to_excel()
            elif event.key() == Qt.Key.Key_F:
                self.search_input.setFocus()
                self.search_input.selectAll()
        super().keyPressEvent(event)