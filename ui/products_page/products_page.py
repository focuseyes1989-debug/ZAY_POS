# ui/products_page/products_page.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QMessageBox, QLabel,
    QFileDialog
)
from PyQt6.QtCore import Qt, pyqtSignal
from models.database import connect_db
from ui.products_page.product_filters import ProductFilters
from ui.products_page.product_card import ProductCards
from ui.products_page.product_table import ProductTable
from ui.products_page.product_service import ProductService
from ui.product_form_dialog import ProductFormDialog
from ui.manage_categories_dialog import ManageCategoriesDialog
from ui.print_barcode_dialog import PrintBarcodeDialog
from utils.language import lang
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from utils.permissions import PermissionManager, Permission
from loguru import logger
from datetime import datetime
import csv


class ProductsPage(QWidget):
    categories_changed = pyqtSignal()  # Signal for category changes
    
    def __init__(self, user_role=None, user_id=None, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.user_id = user_id
        self.current_filter = None
        self.current_page = 1
        self.items_per_page = 50
        self.selected_product_id = None

        self.service = ProductService(self)
        self.setup_ui()
        self.load_categories()
        self.load_products()
        self.update_cards()

        lang.language_changed.connect(self.retranslateUi)
        self.retranslateUi()
        
        # Apply permissions to buttons
        self.apply_permissions()

    def apply_permissions(self):
        """Apply permissions to buttons based on user role"""
        if self.user_id:
            if not PermissionManager.user_has_permission(self.user_id, Permission.EDIT_PRODUCT):
                self.btn_edit.setEnabled(False)
                self.btn_edit.setToolTip("You don't have permission to edit products")
            
            if not PermissionManager.user_has_permission(self.user_id, Permission.DELETE_PRODUCT):
                self.btn_delete.setEnabled(False)
                self.btn_delete.setToolTip("You don't have permission to delete products")
            
            if not PermissionManager.user_has_permission(self.user_id, Permission.ADD_PRODUCT):
                self.btn_add.setEnabled(False)
                self.btn_add.setToolTip("You don't have permission to add products")

    def setup_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)

        # Cards
        self.cards = ProductCards(self)
        main_layout.addWidget(self.cards)

        # Top bar: filters and buttons
        top_layout = QHBoxLayout()
        self.filters = ProductFilters(self)
        self.filters.filter_changed.connect(self.on_filter_changed)
        self.filters.barcode_scanned.connect(self.on_barcode_scanned)

        self.btn_add = QPushButton("Add Item")
        self.btn_edit = QPushButton("Edit")
        self.btn_delete = QPushButton("Delete")
        self.btn_manage_cat = QPushButton("Manage Categories")
        self.btn_print_barcode = QPushButton("Print Barcode")
        
        # Export Buttons
        self.btn_export_list = QPushButton("📊 Export Excel")
        self.btn_export_list.clicked.connect(self.export_products_to_excel)
        
        self.btn_export_price = QPushButton("💰 Export Price List")
        self.btn_export_price.clicked.connect(self.export_price_list_to_excel)
        
        self.btn_export_barcode = QPushButton("📱 Export Barcode Data")
        self.btn_export_barcode.clicked.connect(self.export_barcode_data_to_excel)

        self.btn_add.clicked.connect(self.open_add_dialog)
        self.btn_edit.clicked.connect(self.edit_product)
        self.btn_delete.clicked.connect(self.delete_product)
        self.btn_manage_cat.clicked.connect(self.open_manage_categories)
        self.btn_print_barcode.clicked.connect(self.print_barcode)

        top_layout.addWidget(self.filters)
        top_layout.addWidget(self.btn_add)
        top_layout.addWidget(self.btn_edit)
        top_layout.addWidget(self.btn_delete)
        top_layout.addWidget(self.btn_manage_cat)
        top_layout.addWidget(self.btn_print_barcode)
        top_layout.addWidget(self.btn_export_list)
        top_layout.addWidget(self.btn_export_price)
        top_layout.addWidget(self.btn_export_barcode)
        main_layout.addLayout(top_layout)

        # Product table
        self.table = ProductTable(self)
        self.table.product_selected.connect(self.on_product_selected)
        self.table.service_selected.connect(self.on_service_selected)
        main_layout.addWidget(self.table)

        # Bottom bar
        bottom_layout = QHBoxLayout()
        self.total_label = QLabel("Total Products: 0")
        self.btn_export = QPushButton("Export CSV")
        self.btn_import = QPushButton("Import CSV")
        self.btn_export.clicked.connect(self.export_products)
        self.btn_import.clicked.connect(self.import_products)
        bottom_layout.addWidget(self.total_label)
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.btn_export)
        bottom_layout.addWidget(self.btn_import)
        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)

    def on_card_filter(self, key):
        """Card နှိပ်လိုက်ရင် filter သတ်မှတ်ပြီး products ပြန်တင်မယ်"""
        if key == "total_cost":
            self.current_filter = None
            self.filters.reset()
        else:
            self.current_filter = key
        
        self.current_page = 1
        self.apply_filter()

    def on_filter_changed(self):
        self.current_filter = None
        self.current_page = 1
        self.apply_filter()

    def on_barcode_scanned(self, keyword):
        main_window = self.window()
        if hasattr(main_window, 'sales_page'):
            main_window.switch_to_page(5)
            main_window.sales_page.product_grid.barcode_scanned.emit(keyword)

    def apply_filter(self):
        if self.current_filter == "out_of_stock":
            rows, total = self.service.filter_by_type(
                'out_of_stock', self.current_page, self.items_per_page
            )
        elif self.current_filter == "low_stock":
            rows, total = self.service.filter_by_type(
                'low_stock', self.current_page, self.items_per_page
            )
        elif self.current_filter == "expiring_soon":
            rows, total = self.service.filter_by_type(
                'expiring_soon', self.current_page, self.items_per_page
            )
        elif self.current_filter == "expired":
            rows, total = self.service.filter_by_type(
                'expired', self.current_page, self.items_per_page
            )
        else:
            rows, total = self.service.load_products(
                self.current_page, self.items_per_page,
                self.filters.get_search_text(),
                self.filters.get_category()
            )
        self.table.set_pagination_total(total)
        self.table.populate_table(rows)
        self.update_total_label(total)

    def load_products(self):
        self.apply_filter()

    def load_categories(self):
        self.filters.load_categories()

    def update_cards(self):
        self.cards.update_cards()

    def update_total_label(self, count):
        if lang.get_current() == "my":
            self.total_label.setText(f"စုစုပေါင်းပစ္စည်း: {count}")
        else:
            self.total_label.setText(f"Total Products: {count}")

    def on_product_selected(self, prod_id, name, price, stock):
        self.selected_product_id = prod_id

    def on_service_selected(self, prod_id, name, price):
        self.selected_product_id = prod_id

    def open_add_dialog(self):
        # Check permission
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, Permission.ADD_PRODUCT):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to add products.")
            return
            
        dialog = ProductFormDialog()
        if dialog.exec():
            product_name = dialog.name_input.text()
            self.load_categories()
            self.load_products()
            self.update_cards()
            main_window = self.window()
            if hasattr(main_window, 'inventory_page'):
                main_window.inventory_page.refresh_all()
            if hasattr(main_window, 'current_user'):
                self.service.log_activity(main_window.current_user["id"], main_window.current_user["username"],
                                         "Add Product", f"Product: {product_name}")
            # Emit signal to update sales page categories
            self.categories_changed.emit()
            logger.info("New product added")

    def edit_product(self):
        # Check permission
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, Permission.EDIT_PRODUCT):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to edit products.")
            return
            
        prod_id = self.table.get_selected_product_id()
        if not prod_id:
            QMessageBox.warning(self, "No Selection", "Please select a product first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM products WHERE id=?", (prod_id,))
        row = cursor.fetchone()
        product_name = row[0] if row else "Unknown"
        conn.close()
        dialog = ProductFormDialog(prod_id)
        if dialog.exec():
            self.load_categories()
            self.load_products()
            self.update_cards()
            main_window = self.window()
            if hasattr(main_window, 'inventory_page'):
                main_window.inventory_page.refresh_all()
            if hasattr(main_window, 'current_user'):
                self.service.log_activity(main_window.current_user["id"], main_window.current_user["username"],
                                         "Edit Product", f"Product: {product_name}")
            logger.info(f"Product edited: ID {prod_id}")

    def delete_product(self):
        # Check permission
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, Permission.DELETE_PRODUCT):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to delete products.")
            return
            
        prod_id = self.table.get_selected_product_id()
        if not prod_id:
            QMessageBox.warning(self, "No Selection", "Please select a product first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM products WHERE id=?", (prod_id,))
        row = cursor.fetchone()
        product_name = row[0] if row else "Unknown"
        conn.close()
        reply = QMessageBox.question(self, "Confirm Delete", "Delete this product permanently?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM products WHERE id=?", (prod_id,))
            conn.commit()
            conn.close()
            QMessageBox.information(self, "Deleted", "Product Deleted")
            main_window = self.window()
            if hasattr(main_window, 'current_user'):
                self.service.log_activity(main_window.current_user["id"], main_window.current_user["username"],
                                         "Delete Product", f"Product: {product_name}")
            logger.info(f"Product deleted: ID {prod_id}")
            self.load_categories()
            self.load_products()
            self.update_cards()
            if hasattr(main_window, 'inventory_page'):
                main_window.inventory_page.refresh_all()

    def print_barcode(self):
        prod_id = self.table.get_selected_product_id()
        if not prod_id:
            QMessageBox.warning(self, "No Selection", "Please select a product first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name, barcode FROM products WHERE id=?", (prod_id,))
        row = cursor.fetchone()
        conn.close()
        if row:
            name, barcode = row
            if not barcode:
                QMessageBox.warning(self, "No Barcode", "This product does not have a barcode number.\nPlease edit the product and add a barcode.")
                return
            dialog = PrintBarcodeDialog(prod_id, name, barcode, self)
            dialog.exec()
        else:
            QMessageBox.warning(self, "Error", "Product not found.")

    def open_manage_categories(self):
        dialog = ManageCategoriesDialog()
        # Connect signal to refresh categories in products page
        dialog.categories_changed.connect(self.on_categories_changed)
        if dialog.exec():
            self.load_categories()
            self.load_products()
            self.update_cards()
            # Emit signal to notify main window (which will forward to sales page)
            self.categories_changed.emit()
            logger.info("Categories managed")

    def on_categories_changed(self):
        """Called when categories are added/edited/deleted"""
        self.load_categories()
        self.load_products()
        self.update_cards()
        # Emit signal to update sales page
        self.categories_changed.emit()

    def export_products(self):
        rows = self.table.get_current_rows()
        self.service.export_products(rows, self)

    def import_products(self):
        self.service.import_products(self, self.refresh_after_import)

    def refresh_after_import(self):
        self.load_categories()
        self.load_products()
        self.update_cards()
        main_window = self.window()
        if hasattr(main_window, 'inventory_page'):
            main_window.inventory_page.refresh_all()

    # ========== EXCEL EXPORT FUNCTIONS ==========
    
    def get_all_products_data(self):
        """Get all products data for export"""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, sku, name, category, barcode, price, cost, stock, 
                   low_stock, sold_by, expire_date
            FROM products
            ORDER BY name
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_products_to_excel(self):
        """Export complete product list to Excel"""
        lang_code = lang.get_current()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"product_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Product List" if lang_code != "my" else "ပစ္စည်းစာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            rows = self.get_all_products_data()
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Title
            ws.merge_cells('A1:J1')
            ws['A1'] = "PRODUCT LIST REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Headers
            if lang_code == "my":
                headers = ["SKU", "ပစ္စည်းအမည်", "အမျိုးအစား", "ဘားကုဒ်", 
                          "ရောင်းဈေး", "ကုန်ကျစရိတ်", "ကျန်", 
                          "သတိပေးပမာဏ", "ရောင်းပုံစံ", "သက်တမ်းကုန်ရက်"]
            else:
                headers = ["SKU", "Product Name", "Category", "Barcode", 
                          "Selling Price", "Cost", "Stock", 
                          "Low Stock Alert", "Sold By", "Expiry Date"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            total_stock = 0
            total_stock_value = 0
            total_cost_value = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                pid, sku, name, category, barcode, price, cost, stock, low_stock, sold_by, expire_date = row_data
                
                price_val = float(price) if price else 0
                cost_val = float(cost) if cost else 0
                stock_val = int(stock) if stock else 0
                
                stock_value = price_val * stock_val
                cost_value = cost_val * stock_val
                
                ws.cell(row=row_idx, column=1, value=sku or "")
                ws.cell(row=row_idx, column=2, value=name or "")
                ws.cell(row=row_idx, column=3, value=category or "")
                ws.cell(row=row_idx, column=4, value=barcode or "")
                ws.cell(row=row_idx, column=5, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=6, value=format_money(cost_val, symbol))
                ws.cell(row=row_idx, column=7, value=stock_val)
                ws.cell(row=row_idx, column=8, value=low_stock or 0)
                ws.cell(row=row_idx, column=9, value=sold_by or "Each")
                ws.cell(row=row_idx, column=10, value=expire_date or "")
                
                total_stock += stock_val
                total_stock_value += stock_value
                total_cost_value += cost_value
            
            # Summary rows
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=6, value="TOTAL STOCK VALUE:").font = Font(bold=True)
            ws.cell(row=summary_row, column=7, value=format_money(total_stock_value, symbol))
            ws.cell(row=summary_row + 1, column=6, value="TOTAL COST VALUE:").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=7, value=format_money(total_cost_value, symbol))
            ws.cell(row=summary_row + 2, column=6, value="POTENTIAL PROFIT:").font = Font(bold=True)
            ws.cell(row=summary_row + 2, column=7, value=format_money(total_stock_value - total_cost_value, symbol))
            ws.cell(row=summary_row + 3, column=6, value="TOTAL QUANTITY:").font = Font(bold=True)
            ws.cell(row=summary_row + 3, column=7, value=total_stock)
            
            # Auto adjust columns
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 30
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def export_price_list_to_excel(self):
        """Export price list (only name, SKU, price) to Excel"""
        lang_code = lang.get_current()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"price_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Price List" if lang_code != "my" else "စျေးနှုန်းစာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT sku, name, category, price, sold_by, barcode
                FROM products
                ORDER BY name
            """)
            rows = cursor.fetchall()
            conn.close()
            
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Price List"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "PRICE LIST REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Headers
            if lang_code == "my":
                headers = ["SKU", "ပစ္စည်းအမည်", "အမျိုးအစား", "စျေးနှုန်း", "ရောင်းပုံစံ", "ဘားကုဒ်"]
            else:
                headers = ["SKU", "Product Name", "Category", "Price", "Sold By", "Barcode"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=6):
                sku, name, category, price, sold_by, barcode = row_data
                price_val = float(price) if price else 0
                
                ws.cell(row=row_idx, column=1, value=sku or "")
                ws.cell(row=row_idx, column=2, value=name or "")
                ws.cell(row=row_idx, column=3, value=category or "")
                ws.cell(row=row_idx, column=4, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=5, value=sold_by or "Each")
                ws.cell(row=row_idx, column=6, value=barcode or "")
            
            # Auto adjust columns
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 30
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def export_barcode_data_to_excel(self):
        """Export barcode data for label printing to Excel"""
        lang_code = lang.get_current()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"barcode_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Barcode Data" if lang_code != "my" else "ဘားကုဒ်ဒေတာ ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT sku, name, barcode, price, category
                FROM products
                WHERE barcode IS NOT NULL AND barcode != ''
                ORDER BY name
            """)
            rows = cursor.fetchall()
            conn.close()
            
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcode Data"
            
            # Title
            ws.merge_cells('A1:E1')
            ws['A1'] = "BARCODE DATA REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products with Barcode: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            ws['A4'] = "NOTE: This file can be used for barcode label printing"
            ws['A4'].font = Font(size=9, italic=True, color="7f8c8d")
            
            # Headers
            if lang_code == "my":
                headers = ["SKU", "ပစ္စည်းအမည်", "ဘားကုဒ်", "စျေးနှုန်း", "အမျိုးအစား"]
            else:
                headers = ["SKU", "Product Name", "Barcode", "Price", "Category"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=7):
                sku, name, barcode, price, category = row_data
                price_val = float(price) if price else 0
                
                ws.cell(row=row_idx, column=1, value=sku or "")
                ws.cell(row=row_idx, column=2, value=name or "")
                ws.cell(row=row_idx, column=3, value=barcode or "")
                ws.cell(row=row_idx, column=4, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=5, value=category or "")
            
            # Tips section
            tips_row = len(rows) + 8
            ws.cell(row=tips_row, column=1, value="BARCODE FORMAT TIPS:").font = Font(bold=True, size=11)
            ws.cell(row=tips_row + 1, column=1, value="- Code128: Supports alphanumeric, variable length")
            ws.cell(row=tips_row + 2, column=1, value="- EAN13: 13 digits, for retail products")
            ws.cell(row=tips_row + 3, column=1, value="- UPC-A: 12 digits, for North American products")
            
            # Auto adjust columns
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 30
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def retranslateUi(self):
        self.filters.retranslateUi()
        self.cards.retranslateUi()
        self.table.retranslateUi()

        if lang.get_current() == "my":
            self.btn_add.setText("ပစ္စည်းအသစ်")
            self.btn_edit.setText("ပြင်မည်")
            self.btn_delete.setText("ဖျက်မည်")
            self.btn_manage_cat.setText("အမျိုးအစားများ")
            self.btn_print_barcode.setText("ဘားကုဒ်ထုတ်မည်")
            self.btn_export_list.setText("📊 Excel ထုတ်မည်")
            self.btn_export_price.setText("💰 စျေးနှုန်းစာရင်း")
            self.btn_export_barcode.setText("📱 ဘားကုဒ်ဒေတာ")
            self.btn_export.setText("CSV ထုတ်မည်")
            self.btn_import.setText("CSV သွင်းမည်")
        else:
            self.btn_add.setText("Add Item")
            self.btn_edit.setText("Edit")
            self.btn_delete.setText("Delete")
            self.btn_manage_cat.setText("Manage Categories")
            self.btn_print_barcode.setText("Print Barcode")
            self.btn_export_list.setText("📊 Export Excel")
            self.btn_export_price.setText("💰 Export Price List")
            self.btn_export_barcode.setText("📱 Export Barcode Data")
            self.btn_export.setText("Export CSV")
            self.btn_import.setText("Import CSV")

    def showEvent(self, event):
        self.load_products()
        self.update_cards()
        super().showEvent(event)