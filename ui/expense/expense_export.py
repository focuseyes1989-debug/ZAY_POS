# ui/expense/expense_export.py
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from PyQt6.QtCore import QDate
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from datetime import datetime
import csv


class ExpenseExport:
    """Export functions for expense page"""
    
    @staticmethod
    def get_lang():
        try:
            from models.database import connect_db
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"
    
    @staticmethod
    def export_expense_report(parent, from_date, to_date, category, search_text):
        """Export expense list to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            parent, 
            f"expense_report_{from_date}_to_{to_date}.xlsx",
            "Export Expense Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            lang = ExpenseExport.get_lang()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            query = """
                SELECT expense_no, expense_date, category, description, 
                       amount, payment_method, reference_no, notes
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """
            params = [from_date, to_date]
            
            if search_text:
                query += " AND (LOWER(description) LIKE ? OR LOWER(reference_no) LIKE ?)"
                like = f'%{search_text}%'
                params.extend([like, like])
            
            if category != "All Categories":
                query += " AND category = ?"
                params.append(category)
            
            query += " ORDER BY expense_date DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            total_amount = sum(row[4] for row in rows) if rows else 0
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Expense Report"
            
            # Title
            ws.merge_cells('A1:H1')
            ws['A1'] = "EXPENSE REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Headers
            headers = ["Expense No", "Date", "Category", "Description", 
                      "Amount", "Payment Method", "Reference", "Notes"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=6):
                ws.cell(row=row_idx, column=1, value=row_data[0] or "")
                ws.cell(row=row_idx, column=2, value=row_data[1] or "")
                ws.cell(row=row_idx, column=3, value=row_data[2] or "")
                ws.cell(row=row_idx, column=4, value=row_data[3] or "")
                ws.cell(row=row_idx, column=5, value=format_money(row_data[4], symbol))
                ws.cell(row=row_idx, column=6, value=row_data[5] or "")
                ws.cell(row=row_idx, column=7, value=row_data[6] or "")
                ws.cell(row=row_idx, column=8, value=row_data[7] or "")
            
            # Summary
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=4, value="TOTAL").font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=format_money(total_amount, symbol))
            
            # Adjust columns
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            wb.save(file_path)
            ExcelExporter.show_success_message(parent, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(parent, e)
    
    @staticmethod
    def export_category_report(parent, from_date, to_date):
        """Export category summary to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            parent, 
            f"expense_by_category_{from_date}_to_{to_date}.xlsx",
            "Export Category Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            lang = ExpenseExport.get_lang()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT category, COUNT(*) as count, COALESCE(SUM(amount), 0) as total,
                       COALESCE(AVG(amount), 0) as average, MIN(amount) as min_amount, MAX(amount) as max_amount
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                GROUP BY category
                ORDER BY total DESC
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE expense_date BETWEEN ? AND ?", 
                          (from_date, to_date))
            total_all = cursor.fetchone()[0]
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Category Summary"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = "EXPENSE BY CATEGORY"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"Grand Total: {format_money(total_all, symbol)}"
            
            # Headers
            headers = ["Category", "Count", "Total Amount", "Average", "Min", "Max", "Percentage"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=7):
                category, count, total, avg, min_amt, max_amt = row_data
                percentage = (total / total_all * 100) if total_all > 0 else 0
                
                ws.cell(row=row_idx, column=1, value=category)
                ws.cell(row=row_idx, column=2, value=count)
                ws.cell(row=row_idx, column=3, value=format_money(total, symbol))
                ws.cell(row=row_idx, column=4, value=format_money(avg, symbol))
                ws.cell(row=row_idx, column=5, value=format_money(min_amt, symbol))
                ws.cell(row=row_idx, column=6, value=format_money(max_amt, symbol))
                ws.cell(row=row_idx, column=7, value=f"{percentage:.1f}%")
            
            # Adjust columns
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            wb.save(file_path)
            ExcelExporter.show_success_message(parent, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(parent, e)
    
    @staticmethod
    def export_monthly_report(parent, from_date, to_date):
        """Export monthly summary to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            parent, 
            f"monthly_expense_{from_date}_to_{to_date}.xlsx",
            "Export Monthly Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            lang = ExpenseExport.get_lang()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT strftime('%Y', expense_date) as year, strftime('%m', expense_date) as month,
                       strftime('%Y-%m', expense_date) as year_month, COUNT(*) as count,
                       COALESCE(SUM(amount), 0) as total
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                GROUP BY strftime('%Y-%m', expense_date)
                ORDER BY year_month
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE expense_date BETWEEN ? AND ?", 
                          (from_date, to_date))
            total_all = cursor.fetchone()[0]
            conn.close()
            
            month_names = {
                "01": "January", "02": "February", "03": "March", "04": "April",
                "05": "May", "06": "June", "07": "July", "08": "August",
                "09": "September", "10": "October", "11": "November", "12": "December"
            }
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Monthly Summary"
            
            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = "MONTHLY EXPENSE REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"Total Expenses: {format_money(total_all, symbol)}"
            
            # Headers
            headers = ["Year", "Month", "Transaction Count", "Total Amount"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=7):
                year, month, year_month, count, total = row_data
                month_name = month_names.get(month, month)
                
                ws.cell(row=row_idx, column=1, value=year)
                ws.cell(row=row_idx, column=2, value=f"{month_name} ({month})")
                ws.cell(row=row_idx, column=3, value=count)
                ws.cell(row=row_idx, column=4, value=format_money(total, symbol))
            
            # Adjust columns
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 20
            
            wb.save(file_path)
            ExcelExporter.show_success_message(parent, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(parent, e)