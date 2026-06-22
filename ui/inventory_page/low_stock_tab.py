from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QFileDialog, QHBoxLayout, QMessageBox
from PyQt6.QtCore import Qt
from models.database import connect_db
from ui.widgets.pagination_widget import PaginationWidget
from datetime import datetime
import csv


class LowStockTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        layout = QVBoxLayout()
        
        # Button layout for export
        btn_layout = QHBoxLayout()
        self.btn_export = QPushButton("📊 Export Low Stock Report")
        self.btn_export.clicked.connect(self.export_to_excel)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_export)
        layout.addLayout(btn_layout)
        
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)

    def on_page_changed(self, page: int, page_size: int):
        self.load_data(page, page_size)

    def refresh(self):
        self.load_data()

    def load_data(self, page=1, page_size=50):
        lang = self.get_lang()
        if lang == "my":
            headers = [
                "ပစ္စည်းအမည်", "SKU", "လက်ကျန်", "အနည်းဆုံးပမာဏ", "ပြန်မှာသင့်ပမာဏ",
                "ပေးသွင်းသူ", "နောက်ဆုံးဝယ်ယူရက်", "အခြေအနေ"
            ]
        else:
            headers = [
                "Product Name", "SKU", "Current Qty", "Minimum Qty", "Suggested Reorder Qty",
                "Supplier", "Last Purchase Date", "Status"
            ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM products p
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)
        """)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        cursor.execute("""
            SELECT p.name, p.sku, p.stock, p.low_stock, (COALESCE(p.low_stock, 0) * 2) as suggested,
                   s.name as supplier,
                   (SELECT MAX(created_at) FROM stock_movements WHERE product_id=p.id AND type='in') as last_purchase,
                   CASE WHEN COALESCE(p.stock, 0) = 0 THEN 'Critical' ELSE 'Warning' END as status
            FROM products p
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)
            ORDER BY p.stock ASC
            LIMIT ? OFFSET ?
        """, (page_size, offset))
        rows = cursor.fetchall()
        conn.close()

        self.table.setRowCount(0)
        for row in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            for col, val in enumerate(row):
                self.table.setItem(r, col, QTableWidgetItem(str(val) if val else ""))

    def get_all_low_stock_data(self):
        """Get all low stock data for export"""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.name, p.sku, p.stock, p.low_stock, (COALESCE(p.low_stock, 0) * 2) as suggested,
                   s.name as supplier,
                   (SELECT MAX(created_at) FROM stock_movements WHERE product_id=p.id AND type='in') as last_purchase,
                   CASE WHEN COALESCE(p.stock, 0) = 0 THEN 'Critical' ELSE 'Warning' END as status
            FROM products p
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)
            ORDER BY p.stock ASC
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_to_excel(self):
        """Export low stock report to Excel"""
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"low_stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Low Stock Report" if lang != "my" else "စတော့နည်းနေသောစာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_low_stock_data()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Low Stock Alert"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = "LOW STOCK ALERT REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Products with Low Stock: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Headers
            if lang == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "လက်ကျန်", "အနည်းဆုံးပမာဏ", 
                          "ပြန်မှာသင့်ပမာဏ", "ပေးသွင်းသူ", "အခြေအနေ"]
            else:
                headers = ["Product Name", "SKU", "Current Stock", "Min Stock", 
                          "Suggested Order", "Supplier", "Status"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            critical_count = 0
            warning_count = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                name, sku, stock, low_stock, suggested, supplier, last_purchase, status = row_data
                
                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=stock)
                ws.cell(row=row_idx, column=4, value=low_stock)
                ws.cell(row=row_idx, column=5, value=suggested)
                ws.cell(row=row_idx, column=6, value=supplier or "No Supplier")
                
                status_cell = ws.cell(row=row_idx, column=7, value=status)
                if status == "Critical":
                    status_cell.font = Font(color="FF0000", bold=True)
                    critical_count += 1
                else:
                    status_cell.font = Font(color="FF8C00", bold=True)
                    warning_count += 1
            
            # Summary
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=5, value="SUMMARY").font = Font(bold=True, size=12)
            ws.cell(row=summary_row + 1, column=5, value=f"Critical (Out of Stock): {critical_count}")
            ws.cell(row=summary_row + 2, column=5, value=f"Warning (Low Stock): {warning_count}")
            ws.cell(row=summary_row + 3, column=5, value=f"Total: {len(rows)}")
            
            # Auto adjust columns
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 18
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"