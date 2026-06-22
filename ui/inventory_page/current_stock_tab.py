# ui/inventory_page/current_stock_tab.py
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QMessageBox, QFileDialog, QLabel, QLineEdit, QComboBox
from PyQt6.QtCore import Qt
from models.database import connect_db
from utils.currency import format_money, get_currency_symbol
from utils.translations import tr
from ui.widgets.pagination_widget import PaginationWidget
from ui.inventory_page.product_transaction_history_dialog import ProductTransactionHistoryDialog
from datetime import datetime
import csv


class CurrentStockTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        self.current_page = 1
        self.page_size = 50
        layout = QVBoxLayout()

        # Top button layout
        btn_layout = QHBoxLayout()
        self.btn_stock_in = QPushButton(tr("stock_in"))
        self.btn_stock_in.clicked.connect(self.open_stock_in)
        self.btn_stock_out = QPushButton(tr("stock_out"))
        self.btn_stock_out.clicked.connect(self.open_stock_out)
        self.btn_adjustment = QPushButton(tr("adjustment"))
        self.btn_adjustment.clicked.connect(self.open_adjustment)
        
        # Transfer button
        self.btn_transfer = QPushButton("🔄 Transfer")
        self.btn_transfer.clicked.connect(self.open_transfer)
        
        # Export button
        self.btn_export = QPushButton("📊 Export Current Stock")
        self.btn_export.clicked.connect(self.export_to_excel)
        
        btn_layout.addWidget(self.btn_stock_in)
        btn_layout.addWidget(self.btn_stock_out)
        btn_layout.addWidget(self.btn_adjustment)
        btn_layout.addWidget(self.btn_transfer)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_export)
        layout.addLayout(btn_layout)

        # Filter section
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name, SKU or barcode...")
        self.search_input.textChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_input, 2)
        
        # Category filter
        self.category_filter = QComboBox()
        self.category_filter.addItem("All Categories")
        self.category_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(QLabel("Category:"))
        filter_layout.addWidget(self.category_filter, 1)
        
        # Status filter
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All Status", "In Stock", "Low Stock", "Out of Stock"])
        self.status_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(QLabel("Status:"))
        filter_layout.addWidget(self.status_filter, 1)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.stock_table = QTableWidget()
        self.stock_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.stock_table.verticalHeader().setDefaultSectionSize(40)
        self.stock_table.verticalHeader().setVisible(True)
        self.stock_table.setAlternatingRowColors(True)
        
        header = self.stock_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        layout.addWidget(self.stock_table)

        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        
        # Load categories and initial data
        self.load_categories()
        self.refresh()
        self.retranslateUi()

    def load_categories(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM categories ORDER BY name")
        rows = cursor.fetchall()
        self.category_filter.blockSignals(True)
        current = self.category_filter.currentText()
        self.category_filter.clear()
        self.category_filter.addItem("All Categories")
        for (name,) in rows:
            self.category_filter.addItem(name)
        idx = self.category_filter.findText(current)
        if idx >= 0:
            self.category_filter.setCurrentIndex(idx)
        else:
            self.category_filter.setCurrentIndex(0)
        self.category_filter.blockSignals(False)
        conn.close()

    def on_filter_changed(self):
        self.current_page = 1
        self.load_data()

    def open_stock_in(self):
        from ui.inventory_page.stock_in_dialog import StockInDialog
        dialog = StockInDialog(self)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def open_stock_out(self):
        from ui.inventory_page.stock_out_dialog import StockOutDialog
        dialog = StockOutDialog(self)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def open_adjustment(self):
        from ui.inventory_page.adjustment_dialog import AdjustmentDialog
        dialog = AdjustmentDialog(self)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def open_transfer(self):
        """Open stock transfer dialog"""
        from ui.inventory_page.stock_transfer_dialog import StockTransferDialog
        dialog = StockTransferDialog(self)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def refresh_stock_alerts(self):
        """Refresh stock alerts in main window"""
        main_window = self.window()
        if hasattr(main_window, 'check_stock_alerts'):
            main_window.check_stock_alerts()

    def on_page_changed(self, page: int, page_size: int):
        self.current_page = page
        self.page_size = page_size
        self.load_data()

    def refresh(self):
        self.load_data()
        self.retranslateUi()

    def load_data(self, page=None, page_size=None):
        if page is None:
            page = self.current_page
        if page_size is None:
            page_size = self.page_size
            
        lang = self.get_lang()
        search_text = self.search_input.text().strip().lower()
        category = self.category_filter.currentText()
        status_filter = self.status_filter.currentText()
        use_category = category != "All Categories"
        
        # Define main headers
        if lang == "my":
            main_headers = [
                "ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", "လက်ကျန်",
                "ကုန်ကျစရိတ်", "ရောင်းဈေး", "စုစုပေါင်းတန်ဖိုး", "သတိပေးပမာဏ",
                "အခြေအနေ", "နောက်ဆုံးပြင်ဆင်ချိန်", "နေရာ"
            ]
        else:
            main_headers = [
                "Product Name", "SKU", "Barcode", "Category", "Current Qty",
                "Cost Price", "Selling Price", "Stock Value", "Low Stock Level",
                "Status", "Last Updated", "Location"
            ]
        
        # Add History column
        headers = main_headers + (["မှတ်တမ်း"] if lang == "my" else ["History"])
        self.stock_table.setColumnCount(len(headers))
        self.stock_table.setHorizontalHeaderLabels(headers)
        
        # Configure column resize modes
        header = self.stock_table.horizontalHeader()
        for col in range(len(headers) - 1):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        
        history_col = len(headers) - 1
        header.setSectionResizeMode(history_col, QHeaderView.ResizeMode.Fixed)
        self.stock_table.setColumnWidth(history_col, 100)

        conn = connect_db()
        cursor = conn.cursor()
        
        # Build base query with filters
        base_query = """
            FROM products p
            LEFT JOIN product_locations pl ON p.id = pl.product_id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
        """
        params = []
        
        if use_category:
            base_query += " AND p.category = ?"
            params.append(category)
        
        if search_text:
            like = f'%{search_text}%'
            base_query += " AND (LOWER(p.name) LIKE ? OR LOWER(p.sku) LIKE ? OR LOWER(p.barcode) LIKE ?)"
            params.extend([like, like, like])
        
        # Apply status filter
        if status_filter == "In Stock" or status_filter == "စတော့ရှိပါ":
            base_query += " AND COALESCE(p.stock, 0) > COALESCE(p.low_stock, 0)"
        elif status_filter == "Low Stock" or status_filter == "စတော့နည်းနေပြီ":
            base_query += " AND COALESCE(p.stock, 0) > 0 AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)"
        elif status_filter == "Out of Stock" or status_filter == "ကုန်သွားပြီ":
            base_query += " AND COALESCE(p.stock, 0) = 0"
        
        # Get total count
        count_query = f"""
            SELECT COUNT(DISTINCT p.id) 
            {base_query}
        """
        cursor.execute(count_query, params)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        
        # Query with filters - FIXED: Remove DISTINCT from GROUP_CONCAT
        cursor.execute(f"""
            SELECT 
                p.id, 
                p.name, 
                p.sku, 
                p.barcode, 
                p.category,
                COALESCE(p.stock, 0) as total_stock,
                COALESCE(p.cost, 0) as cost,
                COALESCE(p.price, 0) as price,
                (COALESCE(p.cost, 0) * COALESCE(p.stock, 0)) as stock_value,
                COALESCE(p.low_stock, 0) as low_stock,
                p.sold_by,
                CASE 
                    WHEN p.sold_by = 'Service' THEN 'Service'
                    WHEN COALESCE(p.stock, 0) = 0 THEN 'Out of Stock'
                    WHEN COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0) THEN 'Low Stock'
                    ELSE 'In Stock'
                END as status,
                strftime('%Y-%m-%d %H:%M', p.last_updated) as last_upd,
                GROUP_CONCAT(pl.location || ' (' || pl.quantity || ')', ', ') as locations
            {base_query}
            GROUP BY p.id
            ORDER BY p.name
            LIMIT ? OFFSET ?
        """, params + [page_size, offset])
        rows = cursor.fetchall()
        conn.close()

        self.stock_table.setRowCount(0)
        for row in rows:
            prod_id = row[0]
            name = row[1]
            sku = row[2]
            barcode = row[3]
            category = row[4]
            stock = row[5]
            cost = row[6]
            price = row[7]
            stock_value = row[8]
            low_stock = row[9]
            sold_by = row[10]
            status = row[11]
            last_upd = row[12]
            locations = row[13]  # This is GROUP_CONCAT result
            
            r = self.stock_table.rowCount()
            self.stock_table.insertRow(r)
            
            self.stock_table.setItem(r, 0, QTableWidgetItem(str(name) if name else ""))
            self.stock_table.setItem(r, 1, QTableWidgetItem(str(sku) if sku else ""))
            self.stock_table.setItem(r, 2, QTableWidgetItem(str(barcode) if barcode else ""))
            self.stock_table.setItem(r, 3, QTableWidgetItem(str(category) if category else ""))
            
            stock_item = QTableWidgetItem(str(stock))
            if stock == 0:
                stock_item.setForeground(Qt.GlobalColor.red)
            elif stock <= low_stock:
                stock_item.setForeground(Qt.GlobalColor.darkYellow)
            self.stock_table.setItem(r, 4, stock_item)
            
            self.stock_table.setItem(r, 5, QTableWidgetItem(format_money(cost)))
            self.stock_table.setItem(r, 6, QTableWidgetItem(format_money(price)))
            self.stock_table.setItem(r, 7, QTableWidgetItem(format_money(stock_value)))
            self.stock_table.setItem(r, 8, QTableWidgetItem(str(low_stock)))
            
            status_item = QTableWidgetItem(str(status))
            if status == "Out of Stock" or status == "ကုန်သွားပြီ":
                status_item.setForeground(Qt.GlobalColor.red)
            elif status == "Low Stock" or status == "စတော့နည်းနေပြီ":
                status_item.setForeground(Qt.GlobalColor.darkYellow)
            else:
                status_item.setForeground(Qt.GlobalColor.darkGreen)
            self.stock_table.setItem(r, 9, status_item)
            
            self.stock_table.setItem(r, 10, QTableWidgetItem(str(last_upd) if last_upd else ""))
            
            # Clean up locations string - remove any trailing commas
            locations_str = str(locations) if locations else ""
            # Remove duplicates and clean up
            if locations_str:
                # Split, remove duplicates, and join
                loc_list = [loc.strip() for loc in locations_str.split(',') if loc.strip()]
                seen = set()
                unique_locs = []
                for loc in loc_list:
                    if loc not in seen:
                        seen.add(loc)
                        unique_locs.append(loc)
                locations_str = ', '.join(unique_locs)
            
            self.stock_table.setItem(r, 11, QTableWidgetItem(locations_str))
            
            # History Button
            btn_history = QPushButton("📋 ကြည့်ရန်" if lang == "my" else "📋 View")
            btn_history.setFixedSize(70, 28)
            btn_history.clicked.connect(lambda checked, pid=prod_id, pname=name: self.show_transaction_history(pid, pname))
            self.stock_table.setCellWidget(r, 12, btn_history)

    def show_transaction_history(self, product_id, product_name):
        dialog = ProductTransactionHistoryDialog(product_id, product_name, self)
        dialog.exec()

    def get_all_stock_data(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                p.id, 
                p.name, 
                p.sku, 
                p.barcode, 
                p.category,
                COALESCE(p.stock, 0) as total_stock,
                p.cost, 
                p.price,
                (COALESCE(p.cost, 0) * COALESCE(p.stock, 0)) as stock_value,
                p.low_stock,
                (SELECT GROUP_CONCAT(pl.location || ' (' || pl.quantity || ')', ', ') 
                 FROM product_locations pl
                 WHERE pl.product_id = p.id AND pl.quantity > 0) as locations
            FROM products p
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
            ORDER BY p.name
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_to_excel(self):
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"current_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Current Stock" if lang != "my" else "လက်ရှိစတော့စာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_stock_data()
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Current Stock"
            
            ws.merge_cells('A1:J1')
            ws['A1'] = "CURRENT STOCK REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            if lang == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", 
                          "လက်ကျန်", "ကုန်ကျစရိတ်", "ရောင်းဈေး", "စတော့တန်ဖိုး", 
                          "သတိပေးပမာဏ", "နေရာ"]
            else:
                headers = ["Product Name", "SKU", "Barcode", "Category", 
                          "Current Stock", "Cost Price", "Selling Price", "Stock Value", 
                          "Low Stock Alert", "Location"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            total_stock = 0
            total_stock_value = 0
            total_cost_value = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                # Skip the id column (index 0)
                name, sku, barcode, category, stock, cost, price, stock_value, low_stock, location = row_data[1:]
                
                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=barcode or "")
                ws.cell(row=row_idx, column=4, value=category or "")
                ws.cell(row=row_idx, column=5, value=stock)
                ws.cell(row=row_idx, column=6, value=format_money(cost, symbol))
                ws.cell(row=row_idx, column=7, value=format_money(price, symbol))
                ws.cell(row=row_idx, column=8, value=format_money(stock_value, symbol))
                ws.cell(row=row_idx, column=9, value=low_stock)
                ws.cell(row=row_idx, column=10, value=location or "")
                
                total_stock += stock
                total_stock_value += stock_value
                total_cost_value += cost * stock
            
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=4, value="TOTAL").font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=total_stock)
            ws.cell(row=summary_row, column=7, value=format_money(total_stock_value, symbol))
            ws.cell(row=summary_row + 1, column=4, value="TOTAL COST VALUE").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=5, value=format_money(total_cost_value, symbol))
            ws.cell(row=summary_row + 2, column=4, value="POTENTIAL PROFIT").font = Font(bold=True)
            ws.cell(row=summary_row + 2, column=5, value=format_money(total_stock_value - total_cost_value, symbol))
            
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 20
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def retranslateUi(self):
        self.btn_stock_in.setText(tr("stock_in"))
        self.btn_stock_out.setText(tr("stock_out"))
        self.btn_adjustment.setText(tr("adjustment"))
        
        lang = self.get_lang()
        if lang == "my":
            self.btn_transfer.setText("🔄 လွှဲပြောင်းမည်")
            self.btn_export.setText("📊 စတော့စာရင်းထုတ်မည်")
            self.search_input.setPlaceholderText("ပစ္စည်းအမည် / SKU / ဘားကုဒ်ဖြင့် ရှာရန်...")
            # Update status filter text
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "စတော့ရှိပါ")
            self.status_filter.setItemText(2, "စတော့နည်းနေပြီ")
            self.status_filter.setItemText(3, "ကုန်သွားပြီ")
        else:
            self.btn_transfer.setText("🔄 Transfer")
            self.btn_export.setText("📊 Export Current Stock")
            self.search_input.setPlaceholderText("Search by name, SKU or barcode...")
            self.status_filter.setItemText(0, "All Status")
            self.status_filter.setItemText(1, "In Stock")
            self.status_filter.setItemText(2, "Low Stock")
            self.status_filter.setItemText(3, "Out of Stock")
        
        # Refresh table after language change
        self.load_data()