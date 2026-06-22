# ui/inventory_page/stock_by_location_tab.py
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QPushButton, 
                             QComboBox, QLabel, QLineEdit)
from PyQt6.QtCore import Qt
from models.database import connect_db
from utils.currency import format_money
from ui.widgets.pagination_widget import PaginationWidget
from datetime import datetime


class StockByLocationTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        layout = QVBoxLayout()
        layout.setSpacing(10)

        # Filter section
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        filter_layout.addWidget(QLabel("Location:"))
        self.location_filter = QComboBox()
        self.location_filter.addItem("All Locations")
        self.location_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.location_filter)
        
        filter_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search product by name or SKU...")
        self.search_input.textChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_input)
        
        filter_layout.addStretch()
        
        self.btn_refresh = QPushButton("🔄 Refresh")
        self.btn_refresh.clicked.connect(self.refresh)
        filter_layout.addWidget(self.btn_refresh)
        
        self.btn_export = QPushButton("📊 Export")
        self.btn_export.clicked.connect(self.export_to_excel)
        filter_layout.addWidget(self.btn_export)
        
        layout.addLayout(filter_layout)

        # Stock by Location table
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        self.load_locations()
        self.load_data()

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def load_locations(self):
        """Load locations from product_locations table"""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT location FROM product_locations 
            WHERE location IS NOT NULL AND location != '' 
            ORDER BY location
        """)
        rows = cursor.fetchall()
        
        self.location_filter.blockSignals(True)
        current = self.location_filter.currentText()
        self.location_filter.clear()
        self.location_filter.addItem("All Locations")
        
        for (name,) in rows:
            self.location_filter.addItem(name)
        
        idx = self.location_filter.findText(current)
        if idx >= 0:
            self.location_filter.setCurrentIndex(idx)
        
        self.location_filter.blockSignals(False)
        conn.close()

    def on_filter_changed(self):
        """Refresh when filter changes"""
        self.pagination.set_current_page(1)
        self.load_data()

    def on_page_changed(self, page: int, page_size: int):
        self.load_data(page, page_size)

    def refresh(self):
        self.load_locations()
        self.load_data()

    def load_data(self, page=1, page_size=50):
        lang = self.get_lang()
        
        # Set headers
        if lang == "my":
            headers = ["ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", 
                      "နေရာ", "လက်ကျန်", "ကုန်ကျစရိတ်", "စုစုပေါင်းတန်ဖိုး", 
                      "နောက်ဆုံးပြင်ဆင်ချိန်"]
        else:
            headers = ["Product Name", "SKU", "Barcode", "Category", 
                      "Location", "Quantity", "Cost Price", "Total Value", 
                      "Last Updated"]
        
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        conn = connect_db()
        cursor = conn.cursor()
        
        location_filter = self.location_filter.currentText()
        search_text = self.search_input.text().strip().lower()
        
        # Build query
        count_query = """
            SELECT COUNT(*)
            FROM product_locations pl
            JOIN products p ON pl.product_id = p.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND pl.quantity > 0
        """
        count_params = []
        
        if location_filter != "All Locations":
            count_query += " AND pl.location = ?"
            count_params.append(location_filter)
        
        if search_text:
            count_query += " AND (LOWER(p.name) LIKE ? OR LOWER(p.sku) LIKE ? OR LOWER(p.barcode) LIKE ?)"
            search_pattern = f"%{search_text}%"
            count_params.extend([search_pattern, search_pattern, search_pattern])
        
        cursor.execute(count_query, count_params)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        
        # Main query
        data_query = """
            SELECT 
                p.name,
                p.sku,
                p.barcode,
                p.category,
                pl.location,
                pl.quantity,
                p.cost,
                (COALESCE(p.cost, 0) * pl.quantity) as total_value,
                strftime('%Y-%m-%d %H:%M', pl.last_updated) as last_upd
            FROM product_locations pl
            JOIN products p ON pl.product_id = p.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND pl.quantity > 0
        """
        data_params = []
        
        if location_filter != "All Locations":
            data_query += " AND pl.location = ?"
            data_params.append(location_filter)
        
        if search_text:
            data_query += " AND (LOWER(p.name) LIKE ? OR LOWER(p.sku) LIKE ? OR LOWER(p.barcode) LIKE ?)"
            search_pattern = f"%{search_text}%"
            data_params.extend([search_pattern, search_pattern, search_pattern])
        
        data_query += " ORDER BY pl.location, p.name LIMIT ? OFFSET ?"
        data_params.extend([page_size, offset])
        
        cursor.execute(data_query, data_params)
        rows = cursor.fetchall()
        conn.close()

        self.table.setRowCount(0)
        
        total_quantity = 0
        total_value = 0
        
        for row in rows:
            name, sku, barcode, category, location, quantity, cost, total_value_row, last_upd = row
            
            r = self.table.rowCount()
            self.table.insertRow(r)
            
            self.table.setItem(r, 0, QTableWidgetItem(str(name) if name else ""))
            self.table.setItem(r, 1, QTableWidgetItem(str(sku) if sku else ""))
            self.table.setItem(r, 2, QTableWidgetItem(str(barcode) if barcode else ""))
            self.table.setItem(r, 3, QTableWidgetItem(str(category) if category else ""))
            self.table.setItem(r, 4, QTableWidgetItem(str(location) if location else ""))
            
            qty_item = QTableWidgetItem(str(quantity))
            if quantity <= 0:
                qty_item.setForeground(Qt.GlobalColor.red)
            self.table.setItem(r, 5, qty_item)
            
            self.table.setItem(r, 6, QTableWidgetItem(format_money(cost)))
            self.table.setItem(r, 7, QTableWidgetItem(format_money(total_value_row)))
            self.table.setItem(r, 8, QTableWidgetItem(str(last_upd) if last_upd else ""))
            
            total_quantity += quantity
            total_value += total_value_row
        
        # Add summary row
        if self.table.rowCount() > 0:
            r = self.table.rowCount()
            self.table.insertRow(r)
            
            # Merge cells for summary
            summary_item = QTableWidgetItem("TOTAL" if lang != "my" else "စုစုပေါင်း")
            self.table.setItem(r, 0, summary_item)
            
            self.table.setItem(r, 5, QTableWidgetItem(str(total_quantity)))
            self.table.setItem(r, 7, QTableWidgetItem(format_money(total_value)))

    def get_all_stock_by_location_data(self):
        """Get all data for export"""
        conn = connect_db()
        cursor = conn.cursor()
        
        location_filter = self.location_filter.currentText()
        search_text = self.search_input.text().strip().lower()
        
        query = """
            SELECT 
                p.name,
                p.sku,
                p.barcode,
                p.category,
                pl.location,
                pl.quantity,
                p.cost,
                (COALESCE(p.cost, 0) * pl.quantity) as total_value,
                strftime('%Y-%m-%d %H:%M', pl.last_updated) as last_upd
            FROM product_locations pl
            JOIN products p ON pl.product_id = p.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND pl.quantity > 0
        """
        params = []
        
        if location_filter != "All Locations":
            query += " AND pl.location = ?"
            params.append(location_filter)
        
        if search_text:
            query += " AND (LOWER(p.name) LIKE ? OR LOWER(p.sku) LIKE ? OR LOWER(p.barcode) LIKE ?)"
            search_pattern = f"%{search_text}%"
            params.extend([search_pattern, search_pattern, search_pattern])
        
        query += " ORDER BY pl.location, p.name"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_to_excel(self):
        """Export stock by location to Excel"""
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"stock_by_location_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Stock by Location" if lang != "my" else "နေရာအလိုက်စတော့စာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_stock_by_location_data()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock by Location"
            
            # Title
            ws.merge_cells('A1:I1')
            ws['A1'] = "STOCK BY LOCATION REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Records: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Headers
            if lang == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", 
                          "နေရာ", "လက်ကျန်", "ကုန်ကျစရိတ်", "စုစုပေါင်းတန်ဖိုး", 
                          "နောက်ဆုံးပြင်ဆင်ချိန်"]
            else:
                headers = ["Product Name", "SKU", "Barcode", "Category", 
                          "Location", "Quantity", "Cost Price", "Total Value", 
                          "Last Updated"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            total_qty = 0
            total_value = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                name, sku, barcode, category, location, qty, cost, value_row, last_upd = row_data
                
                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=barcode or "")
                ws.cell(row=row_idx, column=4, value=category or "")
                ws.cell(row=row_idx, column=5, value=location or "")
                ws.cell(row=row_idx, column=6, value=qty)
                ws.cell(row=row_idx, column=7, value=format_money(cost))
                ws.cell(row=row_idx, column=8, value=format_money(value_row))
                ws.cell(row=row_idx, column=9, value=last_upd or "")
                
                total_qty += qty
                total_value += value_row
            
            # Summary
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=4, value="TOTAL" if lang != "my" else "စုစုပေါင်း").font = Font(bold=True)
            ws.cell(row=summary_row, column=6, value=total_qty)
            ws.cell(row=summary_row, column=8, value=format_money(total_value))
            
            # Auto adjust columns
            for col in range(1, 10):
                ws.column_dimensions[chr(64 + col)].width = 18
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def retranslateUi(self):
        self.load_data()