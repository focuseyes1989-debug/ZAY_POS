from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QMessageBox, QHBoxLayout, QFileDialog
from PyQt6.QtCore import Qt
from models.database import connect_db
from utils.currency import format_money
from ui.widgets.pagination_widget import PaginationWidget
from datetime import datetime
import csv


class PurchaseHistoryTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        layout = QVBoxLayout()
        
        # Button layout for export
        btn_layout = QHBoxLayout()
        self.btn_export = QPushButton("📊 Export Purchase History")
        self.btn_export.clicked.connect(self.export_purchase_history)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_export)
        layout.addLayout(btn_layout)
        
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def on_page_changed(self, page: int, page_size: int):
        self.load_data(page, page_size)

    def refresh(self):
        self.load_data()

    def get_all_purchase_data(self):
        """Get all purchase history data for export"""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                po.po_no,
                s.name as supplier_name,
                GROUP_CONCAT(p.name, ' | ') as product_names,
                GROUP_CONCAT(poi.quantity, ' | ') as quantities,
                GROUP_CONCAT(poi.unit_price, ' | ') as unit_prices,
                po.total_amount,
                po.payment_status,
                po.order_date,
                po.received_by,
                po.notes
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN purchase_order_items poi ON po.id = poi.po_id
            LEFT JOIN products p ON poi.product_id = p.id
            GROUP BY po.id
            ORDER BY po.order_date DESC
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_purchase_history(self):
        """Export purchase history to CSV"""
        lang = self.get_lang()
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export Purchase History" if lang != "my" else "ဝယ်ယူမှုမှတ်တမ်း ထုတ်ရန်", 
            f"purchase_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", 
            "CSV Files (*.csv)"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_purchase_data()
            symbol = get_currency_symbol()
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow(["=" * 90])
                writer.writerow(["PURCHASE HISTORY REPORT"] if lang != "my" else ["ဝယ်ယူမှုမှတ်တမ်း အစီရင်ခံစာ"])
                writer.writerow(["=" * 90])
                writer.writerow([])
                writer.writerow(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow(["Total Purchase Orders:", len(rows)])
                writer.writerow([])
                
                # Column Headers
                if lang == "my":
                    writer.writerow(["ဝယ်ယူမှုအမှတ်", "ပေးသွင်းသူ", "ပစ္စည်းများ", 
                                   "အရေအတွက်များ", "တစ်ခုချင်းကုန်ကျစရိတ်",
                                   "စုစုပေါင်းပမာဏ", "ငွေပေးချေမှုအခြေအနေ", "ဝယ်ယူရက်", 
                                   "လက်ခံသူ", "မှတ်ချက်"])
                else:
                    writer.writerow(["Purchase No", "Supplier", "Products", 
                                   "Quantities", "Unit Costs",
                                   "Total Amount", "Payment Status", "Purchase Date", 
                                   "Received By", "Notes"])
                writer.writerow(["-" * 90])
                
                total_amount = 0
                paid_count = 0
                unpaid_count = 0
                partial_count = 0
                
                for row in rows:
                    po_no, supplier, products, quantities, unit_prices, total, payment_status, order_date, received_by, notes = row
                    
                    writer.writerow([
                        po_no or "",
                        supplier or "",
                        (products[:200] + "...") if products and len(products) > 200 else (products or ""),
                        quantities or "",
                        unit_prices or "",
                        format_money(total, symbol),
                        payment_status or "Unpaid",
                        order_date or "",
                        received_by or "",
                        notes or ""
                    ])
                    
                    total_amount += total if total else 0
                    if payment_status == "Paid":
                        paid_count += 1
                    elif payment_status == "Unpaid":
                        unpaid_count += 1
                    else:
                        partial_count += 1
                
                writer.writerow([])
                writer.writerow(["=" * 90])
                writer.writerow(["SUMMARY"])
                writer.writerow(["-" * 50])
                writer.writerow(["Total Purchase Amount:", format_money(total_amount, symbol)])
                writer.writerow(["Paid Orders:", paid_count])
                writer.writerow(["Unpaid Orders:", unpaid_count])
                writer.writerow(["Partial Payments:", partial_count])
                writer.writerow(["=" * 90])
                writer.writerow(["End of Report"])
            
            msg = f"Purchase history exported successfully to:\n{file_path}" if lang != "my" else f"ဝယ်ယူမှုမှတ်တမ်း အောင်မြင်စွာ ထုတ်ယူပြီးပါပြီ:\n{file_path}"
            QMessageBox.information(self, "Export Complete", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")

    def load_data(self, page=1, page_size=50):
        lang = self.get_lang()
        
        # Define headers - removed Discount and Tax
        if lang == "my":
            headers = [
                "ID", "ဝယ်ယူမှုအမှတ်", "ပေးသွင်းသူ", "ပစ္စည်း", "အရေအတွက်", 
                "တစ်ခုချင်းကုန်ကျစရိတ်", "စုစုပေါင်းပမာဏ", "ငွေပေးချေမှုအခြေအနေ", 
                "ဝယ်ယူရက်", "လက်ခံသူ", "မှတ်ချက်", "ပြင်ဆင်", "ဖျက်"
            ]
        else:
            headers = [
                "ID", "Purchase No", "Supplier", "Product", "Quantity", 
                "Unit Cost", "Total Amount", "Payment Status", 
                "Purchase Date", "Received By", "Notes", "Edit", "Delete"
            ]
        
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        # Hide ID column (column 0)
        self.table.setColumnHidden(0, True)

        conn = connect_db()
        cursor = conn.cursor()
        
        # Get total count for pagination
        cursor.execute("SELECT COUNT(*) FROM purchase_orders")
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        
        # Get purchase orders with their items
        cursor.execute("""
            SELECT 
                po.id,
                po.po_no,
                s.name as supplier_name,
                GROUP_CONCAT(p.name, ', ') as product_names,
                GROUP_CONCAT(poi.quantity, ', ') as quantities,
                GROUP_CONCAT(poi.unit_price, ', ') as unit_prices,
                po.total_amount,
                po.payment_status,
                po.order_date,
                po.received_by,
                po.notes
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN purchase_order_items poi ON po.id = poi.po_id
            LEFT JOIN products p ON poi.product_id = p.id
            GROUP BY po.id
            ORDER BY po.order_date DESC
            LIMIT ? OFFSET ?
        """, (page_size, offset))
        
        rows = cursor.fetchall()
        conn.close()

        self.table.setRowCount(0)
        
        for row in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            
            # Column 0: ID (hidden)
            self.table.setItem(r, 0, QTableWidgetItem(str(row[0])))
            
            # Column 1: Purchase No
            self.table.setItem(r, 1, QTableWidgetItem(str(row[1]) if row[1] else ""))
            
            # Column 2: Supplier
            self.table.setItem(r, 2, QTableWidgetItem(str(row[2]) if row[2] else ""))
            
            # Column 3: Product (combined)
            self.table.setItem(r, 3, QTableWidgetItem(str(row[3]) if row[3] else ""))
            
            # Column 4: Quantity (combined)
            self.table.setItem(r, 4, QTableWidgetItem(str(row[4]) if row[4] else ""))
            
            # Column 5: Unit Cost (combined)
            self.table.setItem(r, 5, QTableWidgetItem(str(row[5]) if row[5] else ""))
            
            # Column 6: Total Amount
            total_amt = row[6] if row[6] else 0
            self.table.setItem(r, 6, QTableWidgetItem(format_money(total_amt)))
            
            # Column 7: Payment Status
            payment_status = row[7] if row[7] else "Unpaid"
            status_item = QTableWidgetItem(payment_status)
            if payment_status == "Paid":
                status_item.setForeground(Qt.GlobalColor.darkGreen)
            elif payment_status == "Unpaid":
                status_item.setForeground(Qt.GlobalColor.darkRed)
            else:  # Partial
                status_item.setForeground(Qt.GlobalColor.darkYellow)
            self.table.setItem(r, 7, status_item)
            
            # Column 8: Purchase Date
            self.table.setItem(r, 8, QTableWidgetItem(str(row[8]) if row[8] else ""))
            
            # Column 9: Received By
            self.table.setItem(r, 9, QTableWidgetItem(str(row[9]) if row[9] else ""))
            
            # Column 10: Notes
            self.table.setItem(r, 10, QTableWidgetItem(str(row[10]) if row[10] else ""))
            
            # Column 11: Edit Button
            po_id = row[0]
            btn_edit = QPushButton("ပြင်ဆင်" if lang == "my" else "Edit")
            btn_edit.clicked.connect(lambda _, pid=po_id: self.edit_purchase_order(pid))
            self.table.setCellWidget(r, 11, btn_edit)
            
            # Column 12: Delete Button
            btn_delete = QPushButton("ဖျက်" if lang == "my" else "Delete")
            btn_delete.clicked.connect(lambda _, pid=po_id: self.delete_purchase_order(pid))
            self.table.setCellWidget(r, 12, btn_delete)

    def edit_purchase_order(self, po_id):
        from ui.inventory_page.purchase_order_dialog import PurchaseOrderEditDialog
        dialog = PurchaseOrderEditDialog(po_id, self)
        if dialog.exec():
            self.load_data()

    def delete_purchase_order(self, po_id):
        lang = self.get_lang()
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT po_no FROM purchase_orders WHERE id = ?", (po_id,))
        row = cursor.fetchone()
        conn.close()
        
        if not row:
            QMessageBox.warning(self, "Error", "Purchase order not found.")
            return
        
        po_no = row[0]
        
        confirm_text = (
            f"Are you sure you want to delete purchase order #{po_no}?\n\n"
            "This will remove the purchase order record from history.\n"
            "⚠️ Stock movements will NOT be reversed.\n"
            "This action cannot be undone."
            if lang != "my" else
            f"ဝယ်ယူမှုအမှာစာ #{po_no} ကို ဖျက်မည်လား?\n\n"
            "ဝယ်ယူမှုမှတ်တမ်းကို ဖျက်ပစ်မည်။\n"
            "⚠️ စတော့လှုပ်ရှားမှုများကို ပြန်မဖြည့်ပေးပါ။\n"
            "ဤလုပ်ဆောင်ချက်ကို နောက်ပြန်မဆုတ်နိုင်ပါ။"
        )
        
        reply = QMessageBox.question(self, "Confirm Delete", confirm_text,
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply != QMessageBox.StandardButton.Yes:
            return

        conn = connect_db()
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM purchase_order_items WHERE po_id = ?", (po_id,))
            cursor.execute("DELETE FROM purchase_orders WHERE id = ?", (po_id,))
            conn.commit()
            QMessageBox.information(self, "Deleted", "Purchase order deleted successfully.")
            self.load_data()
        except Exception as e:
            conn.rollback()
            QMessageBox.critical(self, "Error", f"Failed to delete: {e}")
        finally:
            conn.close()
    
    def retranslateUi(self):
        lang = self.get_lang()
        self.load_data()
    
    def export_to_excel(self):
        """Export purchase history to Excel"""
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"purchase_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Purchase History" if lang != "my" else "ဝယ်ယူမှုမှတ်တမ်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_purchase_data()
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Purchase History"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = "PURCHASE HISTORY REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Purchase Orders: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Headers
            if lang == "my":
                headers = ["ဝယ်ယူမှုအမှတ်", "ပေးသွင်းသူ", "ရက်စွဲ", 
                          "စုစုပေါင်းပမာဏ", "ငွေပေးချေမှုအခြေအနေ", "လက်ခံသူ", "မှတ်ချက်"]
            else:
                headers = ["PO Number", "Supplier", "Date", "Total Amount", "Payment Status", "Received By", "Notes"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            total_amount = 0
            paid_count = 0
            unpaid_count = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                po_no, supplier, order_date, amount, payment_status, received_by, notes = row_data
                # ... (rest of the code similar to previous inventory_page.py)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
