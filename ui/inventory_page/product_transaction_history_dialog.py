# ui/inventory_page/product_transaction_history_dialog.py
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget,
    QTableWidgetItem, QHeaderView, QPushButton, QDateEdit,
    QComboBox, QMessageBox, QFileDialog
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QIcon, QColor
from models.database import connect_db
from utils.currency import format_money
from utils.excel_exporter import ExcelExporter
from ui.widgets.pagination_widget import PaginationWidget
from datetime import datetime


class ProductTransactionHistoryDialog(QDialog):
    def __init__(self, product_id, product_name, parent=None):
        super().__init__(parent)
        self.product_id = product_id
        self.product_name = product_name
        self.current_page = 1
        self.page_size = 50
        self.setWindowTitle(f"Transaction History - {product_name}")
        self.setMinimumSize(1100, 550)
        self.setWindowIcon(QIcon("assets/icons/zaypos.png"))
        self.setModal(True)

        layout = QVBoxLayout()
        layout.setSpacing(10)

        # Header info
        info_layout = QHBoxLayout()
        info_label = QLabel(f"<b>Product:</b> {product_name}")
        info_label.setStyleSheet("font-size: 12pt;")
        info_layout.addWidget(info_label)
        info_layout.addStretch()
        
        self.stock_label = QLabel()
        self.stock_label.setStyleSheet("font-size: 11pt; font-weight: bold;")
        info_layout.addWidget(self.stock_label)
        
        layout.addLayout(info_layout)

        # Filter section
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        filter_layout.addWidget(QLabel("From:"))
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addMonths(-3))
        filter_layout.addWidget(self.from_date)
        
        filter_layout.addWidget(QLabel("To:"))
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        filter_layout.addWidget(self.to_date)
        
        filter_layout.addWidget(QLabel("Action:"))
        self.action_filter = QComboBox()
        self.action_filter.addItems(["All", "Stock In", "Stock Out", "Adjustment", "Sale"])
        self.action_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.action_filter)
        
        self.btn_refresh = QPushButton("Refresh")
        self.btn_refresh.clicked.connect(self.load_history)
        filter_layout.addWidget(self.btn_refresh)
        
        self.btn_export = QPushButton("📊 Export Excel")
        self.btn_export.clicked.connect(self.export_to_excel)
        filter_layout.addWidget(self.btn_export)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Transaction table
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["Date", "Action", "Qty Before", "Qty After", "Quantity", "Location", "User", "Remark"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        # Close button removed - user can close with X button or Esc key

        self.setLayout(layout)
        self.load_current_stock()
        self.load_history()

    def load_current_stock(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT stock FROM products WHERE id = ?", (self.product_id,))
        row = cursor.fetchone()
        conn.close()
        if row:
            stock = row[0] if row[0] else 0
            lang = self.get_lang()
            if lang == "my":
                self.stock_label.setText(f"လက်ကျန်စတော့: {stock}")
            else:
                self.stock_label.setText(f"Current Stock: {stock}")
            if stock <= 0:
                self.stock_label.setStyleSheet("font-size: 11pt; color: #e74c3c; font-weight: bold;")
            else:
                self.stock_label.setStyleSheet("font-size: 11pt; color: #2ecc71; font-weight: bold;")

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def on_filter_changed(self):
        self.current_page = 1
        self.load_history()

    def on_page_changed(self, page, page_size):
        self.current_page = page
        self.page_size = page_size
        self.load_history()

    def load_history(self):
        from_date = self.from_date.date().toString("yyyy-MM-dd")
        to_date = self.to_date.date().toString("yyyy-MM-dd")
        action = self.action_filter.currentText()
        lang = self.get_lang()
        
        # Update headers based on language
        if lang == "my":
            headers = ["ရက်စွဲ", "လုပ်ဆောင်ချက်", "မပြောင်းမီ", "ပြောင်းပြီး", "ပြောင်းလဲမှု", "နေရာ", "အသုံးပြုသူ", "မှတ်ချက်"]
        else:
            headers = ["Date", "Action", "Qty Before", "Qty After", "Quantity", "Location", "User", "Remark"]
        self.table.setHorizontalHeaderLabels(headers)
        
        conn = connect_db()
        cursor = conn.cursor()
        
        # Check if location column exists
        cursor.execute("PRAGMA table_info(stock_movements)")
        columns = [col[1] for col in cursor.fetchall()]
        has_location = 'location' in columns
        
        action_map = {
            "Stock In": "in",
            "Stock Out": "out",
            "Adjustment": "adjustment",
            "Sale": "sale"
        }
        
        count_query = """
            SELECT COUNT(*)
            FROM stock_movements sm
            WHERE sm.product_id = ? 
              AND date(sm.created_at) BETWEEN ? AND ?
        """
        count_params = [self.product_id, from_date, to_date]
        
        if action != "All":
            db_action = action_map.get(action, action.lower())
            count_query += " AND sm.type = ?"
            count_params.append(db_action)
        
        cursor.execute(count_query, count_params)
        result = cursor.fetchone()
        total_items = result[0] if result else 0
        
        if total_items is None:
            total_items = 0
        else:
            total_items = int(total_items)
        
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        # Build query based on whether location column exists
        if has_location:
            data_query = """
                SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                       sm.quantity, sm.created_by, sm.reason, sm.notes,
                       sm.location
                FROM stock_movements sm
                WHERE sm.product_id = ? 
                  AND date(sm.created_at) BETWEEN ? AND ?
            """
        else:
            data_query = """
                SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                       sm.quantity, sm.created_by, sm.reason, sm.notes,
                       NULL as location
                FROM stock_movements sm
                WHERE sm.product_id = ? 
                  AND date(sm.created_at) BETWEEN ? AND ?
            """
        
        data_params = [self.product_id, from_date, to_date]
        
        if action != "All":
            db_action = action_map.get(action, action.lower())
            data_query += " AND sm.type = ?"
            data_params.append(db_action)
        
        data_query += " ORDER BY sm.created_at DESC LIMIT ? OFFSET ?"
        
        offset = (self.current_page - 1) * self.page_size
        cursor.execute(data_query, data_params + [self.page_size, offset])
        rows = cursor.fetchall()
        conn.close()
        
        self.table.setRowCount(0)
        for row in rows:
            if has_location:
                created_at, action_type, old_stock, new_stock, qty, user, reason, notes, location = row
            else:
                created_at, action_type, old_stock, new_stock, qty, user, reason, notes = row
                location = None
            
            if lang == "my":
                action_display = {
                    "in": "စတော့ဝင်",
                    "out": "စတော့ထွက်",
                    "adjustment": "ပြင်ဆင်ချက်",
                    "sale": "ရောင်းချမှု"
                }.get(action_type, action_type)
            else:
                action_display = {
                    "in": "Stock In",
                    "out": "Stock Out",
                    "adjustment": "Adjustment",
                    "sale": "Sale"
                }.get(action_type, action_type)
            
            date_str = created_at[:16] if created_at else ""
            
            if qty is not None:
                qty_display = abs(qty)
                if action_type in ["out", "sale"]:
                    qty_display = f"-{abs(qty)}"
                elif action_type == "adjustment":
                    if new_stock > old_stock:
                        qty_display = f"+{new_stock - old_stock}"
                    else:
                        qty_display = f"{new_stock - old_stock}"
                else:
                    qty_display = f"+{abs(qty)}"
            else:
                qty_display = ""
            
            remark = reason if reason else (notes if notes else "")
            location_display = location if location else "-"
            
            row_num = self.table.rowCount()
            self.table.insertRow(row_num)
            self.table.setItem(row_num, 0, QTableWidgetItem(date_str))
            
            action_item = QTableWidgetItem(action_display)
            if action_type in ["in"]:
                action_item.setForeground(QColor(46, 204, 113))
            elif action_type in ["out", "sale"]:
                action_item.setForeground(QColor(231, 76, 60))
            else:
                action_item.setForeground(QColor(241, 196, 15))
            self.table.setItem(row_num, 1, action_item)
            
            self.table.setItem(row_num, 2, QTableWidgetItem(str(old_stock) if old_stock is not None else ""))
            self.table.setItem(row_num, 3, QTableWidgetItem(str(new_stock) if new_stock is not None else ""))
            self.table.setItem(row_num, 4, QTableWidgetItem(str(qty_display)))
            self.table.setItem(row_num, 5, QTableWidgetItem(str(location_display)))
            self.table.setItem(row_num, 6, QTableWidgetItem(user or ""))
            self.table.setItem(row_num, 7, QTableWidgetItem(remark or ""))

    def export_to_excel(self):
        from_date = self.from_date.date().toString("yyyy-MM-dd")
        to_date = self.to_date.date().toString("yyyy-MM-dd")
        action = self.action_filter.currentText()
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"transaction_history_{self.product_name}_{from_date}_to_{to_date}.xlsx",
            "Export Transaction History" if lang != "my" else "ငွေပေးချေမှုမှတ်တမ်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Check if location column exists
            cursor.execute("PRAGMA table_info(stock_movements)")
            columns = [col[1] for col in cursor.fetchall()]
            has_location = 'location' in columns
            
            action_map = {
                "Stock In": "in",
                "Stock Out": "out",
                "Adjustment": "adjustment",
                "Sale": "sale"
            }
            
            if has_location:
                query = """
                    SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                           sm.quantity, sm.created_by, sm.reason, sm.notes,
                           sm.location
                    FROM stock_movements sm
                    WHERE sm.product_id = ? 
                      AND date(sm.created_at) BETWEEN ? AND ?
                """
            else:
                query = """
                    SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                           sm.quantity, sm.created_by, sm.reason, sm.notes,
                           NULL as location
                    FROM stock_movements sm
                    WHERE sm.product_id = ? 
                      AND date(sm.created_at) BETWEEN ? AND ?
                """
            
            params = [self.product_id, from_date, to_date]
            
            if action != "All":
                db_action = action_map.get(action, action.lower())
                query += " AND sm.type = ?"
                params.append(db_action)
            
            query += " ORDER BY sm.created_at DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Transaction History"
            
            ws.merge_cells('A1:H1')
            ws['A1'] = f"TRANSACTION HISTORY - {self.product_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            if lang == "my":
                headers = ["ရက်စွဲ", "လုပ်ဆောင်ချက်", "မပြောင်းမီ", "ပြောင်းပြီး", "ပြောင်းလဲမှု", "နေရာ", "အသုံးပြုသူ", "မှတ်ချက်"]
            else:
                headers = ["Date", "Action", "Qty Before", "Qty After", "Quantity", "Location", "User", "Remark"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=6):
                if has_location:
                    created_at, action_type, old_stock, new_stock, qty, user, reason, notes, location = row_data
                else:
                    created_at, action_type, old_stock, new_stock, qty, user, reason, notes = row_data
                    location = None
                
                if lang == "my":
                    action_display = {
                        "in": "စတော့ဝင်",
                        "out": "စတော့ထွက်",
                        "adjustment": "ပြင်ဆင်ချက်",
                        "sale": "ရောင်းချမှု"
                    }.get(action_type, action_type)
                else:
                    action_display = {
                        "in": "Stock In",
                        "out": "Stock Out",
                        "adjustment": "Adjustment",
                        "sale": "Sale"
                    }.get(action_type, action_type)
                
                ws.cell(row=row_idx, column=1, value=created_at[:16] if created_at else "")
                ws.cell(row=row_idx, column=2, value=action_display)
                ws.cell(row=row_idx, column=3, value=old_stock if old_stock is not None else "")
                ws.cell(row=row_idx, column=4, value=new_stock if new_stock is not None else "")
                
                if qty is not None:
                    if action_type in ["out", "sale"]:
                        ws.cell(row=row_idx, column=5, value=f"-{abs(qty)}")
                    elif action_type == "adjustment":
                        if new_stock > old_stock:
                            ws.cell(row=row_idx, column=5, value=f"+{new_stock - old_stock}")
                        else:
                            ws.cell(row=row_idx, column=5, value=f"{new_stock - old_stock}")
                    else:
                        ws.cell(row=row_idx, column=5, value=f"+{abs(qty)}")
                else:
                    ws.cell(row=row_idx, column=5, value="")
                
                ws.cell(row=row_idx, column=6, value=location if location else "-")
                ws.cell(row=row_idx, column=7, value=user or "")
                
                remark = reason if reason else (notes if notes else "")
                ws.cell(row=row_idx, column=8, value=remark or "")
            
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['H'].width = 25
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)