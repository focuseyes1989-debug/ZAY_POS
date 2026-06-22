# ui/sales_summary/sales_summary_page.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QDateEdit, QTabWidget, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt, QDate
from ui.sales_summary.base_sales_summary import BaseSalesSummary
from ui.sales_summary.top_items_tab import TopItemsTab
from ui.sales_summary.items_tab import ItemsTab
from ui.sales_summary.categories_tab import CategoriesTab
from ui.sales_summary.payment_tab import PaymentTab
from utils.language import lang
from utils.excel_exporter import ExcelExporter
from datetime import datetime


class SalesSummaryPage(BaseSalesSummary):
    def __init__(self):
        super().__init__()
        main_layout = QVBoxLayout()

        # Filter row
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.btn_today = QPushButton("Today")
        self.btn_this_week = QPushButton("This Week")
        self.btn_this_month = QPushButton("This Month")
        self.btn_last_month = QPushButton("Last Month")
        self.btn_this_year = QPushButton("This Year")

        self.btn_today.clicked.connect(self.set_today_range)
        self.btn_this_week.clicked.connect(self.set_this_week_range)
        self.btn_this_month.clicked.connect(self.set_this_month_range)
        self.btn_last_month.clicked.connect(self.set_last_month_range)
        self.btn_this_year.clicked.connect(self.set_this_year_range)

        filter_layout.addWidget(self.btn_today)
        filter_layout.addWidget(self.btn_this_week)
        filter_layout.addWidget(self.btn_this_month)
        filter_layout.addWidget(self.btn_last_month)
        filter_layout.addWidget(self.btn_this_year)
        filter_layout.addStretch()

        self.from_label = QLabel("From:")
        filter_layout.addWidget(self.from_label)
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))
        self.from_date.dateChanged.connect(self.load_all_tabs)
        filter_layout.addWidget(self.from_date)

        self.to_label = QLabel("To:")
        filter_layout.addWidget(self.to_label)
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        self.to_date.dateChanged.connect(self.load_all_tabs)
        filter_layout.addWidget(self.to_date)

        self.btn_refresh = QPushButton("Refresh")
        self.btn_refresh.clicked.connect(self.load_all_tabs)
        filter_layout.addWidget(self.btn_refresh)

        self.btn_export_excel = QPushButton("📊 Export Excel")
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        filter_layout.addWidget(self.btn_export_excel)

        filter_layout.addStretch()
        main_layout.addLayout(filter_layout)

        # Tabs
        self.tabs = QTabWidget()
        self.tab_names = {
            0: "Top 20 Sales by Item",
            1: "Sales by Item",
            2: "Sales by Category",
            3: "Sales by Payment Type"
        }
        
        # Create tabs
        self.top_items_tab = TopItemsTab(self)
        self.tabs.addTab(self.top_items_tab, self.tab_names[0])
        
        self.items_tab = ItemsTab(self)
        self.tabs.addTab(self.items_tab, self.tab_names[1])
        
        self.categories_tab = CategoriesTab(self)
        self.tabs.addTab(self.categories_tab, self.tab_names[2])
        
        self.payment_tab = PaymentTab(self)
        self.tabs.addTab(self.payment_tab, self.tab_names[3])

        main_layout.addWidget(self.tabs)
        self.setLayout(main_layout)

        lang.language_changed.connect(self.retranslateUi)
        self.load_all_tabs()
        self.retranslateUi()

    def showEvent(self, event):
        self.load_all_tabs()
        super().showEvent(event)

    def load_all_tabs(self):
        from_date, to_date = self.get_date_range()
        lang_code = self.get_lang()
        
        self.top_items_tab.load(from_date, to_date, lang_code)
        self.items_tab.load(from_date, to_date)
        self.categories_tab.load(from_date, to_date)
        self.payment_tab.load(from_date, to_date, lang_code)

    def retranslateUi(self):
        lang_code = self.get_lang()
        
        if lang_code == "my":
            self.btn_today.setText("ယနေ့")
            self.btn_this_week.setText("ဤတစ်ပတ်")
            self.btn_this_month.setText("ဤလ")
            self.btn_last_month.setText("ပြီးခဲ့သည့်လ")
            self.btn_this_year.setText("ဤနှစ်")
            self.from_label.setText("မှ:")
            self.to_label.setText("အထိ:")
            self.btn_refresh.setText("ပြန်လည်")
            self.btn_export_excel.setText("📊 Excel ထုတ်မည်")

            tab_titles_my = {
                0: "ထိပ်ဆုံးရောင်းအားရှိပစ္စည်း ၂၀",
                1: "ပစ္စည်းအလိုက်ရောင်းအား",
                2: "အမျိုးအစားအလိုက်ရောင်းအား",
                3: "ငွေပေးချေမှုအလိုက်ရောင်းအား"
            }
            for idx, title in tab_titles_my.items():
                self.tabs.setTabText(idx, title)
        else:
            self.btn_today.setText("Today")
            self.btn_this_week.setText("This Week")
            self.btn_this_month.setText("This Month")
            self.btn_last_month.setText("Last Month")
            self.btn_this_year.setText("This Year")
            self.from_label.setText("From:")
            self.to_label.setText("To:")
            self.btn_refresh.setText("Refresh")
            self.btn_export_excel.setText("📊 Export Excel")

            for idx, title in self.tab_names.items():
                self.tabs.setTabText(idx, title)
        
        # Retranslate tabs
        self.top_items_tab.retranslateUi()
        self.items_tab.retranslateUi()
        self.categories_tab.retranslateUi()
        
        self.load_all_tabs()

    # Export methods remain the same...
    def export_to_excel(self):
        """Export all sales summary tabs to Excel (.xlsx)"""
        from_date, to_date = self.get_date_range()
        symbol = self.get_currency_symbol()
        lang_code = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"sales_summary_{from_date}_to_{to_date}.xlsx",
            "Export Sales Summary" if lang_code != "my" else "ရောင်းအားအကျဉ်းချုပ် ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            
            # Tab 1: Top Products
            ws1 = wb.active
            ws1.title = "Top Products"
            self._export_top_products(ws1, from_date, to_date, symbol, lang_code)
            
            # Tab 2: Products Detail
            ws2 = wb.create_sheet("Products Detail")
            self._export_products_detail(ws2, from_date, to_date, symbol, lang_code)
            
            # Tab 3: Categories
            ws3 = wb.create_sheet("Categories")
            self._export_categories(ws3, from_date, to_date, symbol, lang_code)
            
            # Tab 4: Payment Types
            ws4 = wb.create_sheet("Payment Types")
            self._export_payment(ws4, from_date, to_date, symbol, lang_code)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def _export_top_products(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sale_items.product_name, COALESCE(SUM(sale_items.total), 0) as total_sales
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY sale_items.product_name
            ORDER BY total_sales DESC
            LIMIT 20
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "TOP 20 SALES BY ITEM"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        headers = ["ပစ္စည်းအမည်", "စုစုပေါင်းရောင်းအား"] if lang_code == "my" else ["Product Name", "Total Sales"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, (name, total) in enumerate(rows, start=6):
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=format_money(total, symbol))
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _export_products_detail(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sale_items.product_name,
                   COALESCE(products.category, 'Uncategorized') as category,
                   COALESCE(SUM(sale_items.qty), 0) as total_qty,
                   COALESCE(SUM(sale_items.total), 0) as net_sales,
                   COALESCE(SUM(products.cost * sale_items.qty), 0) as cogs,
                   COALESCE(SUM(sale_items.total) - SUM(products.cost * sale_items.qty), 0) as gross_profit
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            LEFT JOIN products ON sale_items.product_name = products.name
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY sale_items.product_name
            ORDER BY sale_items.product_name
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "SALES BY PRODUCT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["ပစ္စည်းအမည်", "အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", 
                      "အသားတင်ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Product Name", "Category", "Items Sold", "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_qty = total_sales = total_cogs = total_profit = 0
        for row_idx, row_data in enumerate(rows, start=6):
            name, category, qty, sales, cogs, profit = row_data
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=category)
            ws.cell(row=row_idx, column=3, value=qty)
            ws.cell(row=row_idx, column=4, value=format_money(sales, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=6, value=format_money(profit, symbol))
            total_qty += qty; total_sales += sales; total_cogs += cogs; total_profit += profit
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=total_qty)
        ws.cell(row=summary_row, column=4, value=format_money(total_sales, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=6, value=format_money(total_profit, symbol))
        
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_categories(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(products.category, 'Uncategorized') as category,
                   COALESCE(SUM(sale_items.qty), 0) as items_sold,
                   COALESCE(SUM(sale_items.total), 0) as net_sales,
                   COALESCE(SUM(products.cost * sale_items.qty), 0) as cogs
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            LEFT JOIN products ON sale_items.product_name = products.name
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY products.category
            ORDER BY net_sales DESC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:E1')
        ws['A1'] = "SALES BY CATEGORY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", "စုစုပေါင်းရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Category", "Items Sold", "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_items = total_sales = total_cogs = total_profit = 0
        for row_idx, row_data in enumerate(rows, start=6):
            category, items, sales, cogs = row_data
            profit = sales - cogs
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=items)
            ws.cell(row=row_idx, column=3, value=format_money(sales, symbol))
            ws.cell(row=row_idx, column=4, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(profit, symbol))
            total_items += items; total_sales += sales; total_cogs += cogs; total_profit += profit
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_items)
        ws.cell(row=summary_row, column=3, value=format_money(total_sales, symbol))
        ws.cell(row=summary_row, column=4, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_profit, symbol))
        
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_payment(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(payment_type, 'Other') as payment_type,
                   COUNT(*) as transaction_count,
                   COALESCE(SUM(total), 0) as total_amount
            FROM sales
            WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            GROUP BY payment_type
            ORDER BY payment_type
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "SALES BY PAYMENT TYPE"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["ငွေပေးချေမှုအမျိုးအစား", "ငွေပေးချေမှုအရေအတွက်", "ငွေပေးချေမှုပမာဏ"]
        else:
            headers = ["Payment Type", "Transaction Count", "Amount"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_count = total_amount = 0
        for row_idx, row_data in enumerate(rows, start=6):
            ptype, count, amount = row_data
            ws.cell(row=row_idx, column=1, value=ptype)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=format_money(amount, symbol))
            total_count += count; total_amount += amount
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_count)
        ws.cell(row=summary_row, column=3, value=format_money(total_amount, symbol))
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20