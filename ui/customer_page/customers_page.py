# ui/customer_page/customers_page.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit,
    QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView,
    QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from ui.widgets.pagination_widget import PaginationWidget
from utils.language import lang
from utils.activity_logger import log_activity
from ui.customer_page.credit_sale_dialog import CreditSaleDialog
from ui.customer_page.credit_payment_dialog import CreditPaymentDialog
from ui.customer_page.customer_ledger_dialog import CustomerLedgerDialog
from ui.customer_page.outstanding_report_dialog import OutstandingReportDialog
from ui.customer_page.add_edit_customer_dialog import AddEditCustomerDialog
from utils.excel_exporter import ExcelExporter
from datetime import datetime
import csv


class CustomersPage(QWidget):
    def __init__(self, user_role=None, parent=None):
        super().__init__(parent)
        self.user_role = user_role
        self.selected_customer_id = None
        self.current_language = lang.get_current()

        layout = QVBoxLayout()
        layout.setSpacing(10)

        # Search
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name, phone, email...")
        self.search_input.textChanged.connect(self.reset_and_search)
        layout.addWidget(self.search_input)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(5)
        
        self.btn_add = QPushButton()
        self.btn_edit = QPushButton()
        self.btn_delete = QPushButton()
        self.btn_credit_sale = QPushButton()
        self.btn_payment = QPushButton()
        self.btn_ledger = QPushButton()
        self.btn_outstanding_report = QPushButton()
        self.btn_export_excel = QPushButton("📊 Export Excel")
        
        self.btn_add.clicked.connect(self.add_customer)
        self.btn_edit.clicked.connect(self.edit_customer)
        self.btn_delete.clicked.connect(self.delete_customer)
        self.btn_credit_sale.clicked.connect(self.credit_sale)
        self.btn_payment.clicked.connect(self.payment_collection)
        self.btn_ledger.clicked.connect(self.show_ledger)
        self.btn_outstanding_report.clicked.connect(self.show_outstanding_report)
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        
        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_edit)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_credit_sale)
        btn_layout.addWidget(self.btn_payment)
        btn_layout.addWidget(self.btn_ledger)
        btn_layout.addWidget(self.btn_outstanding_report)
        btn_layout.addWidget(self.btn_export_excel)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setColumnHidden(0, True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.cellClicked.connect(self.select_customer)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for col in range(2, 10):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)

        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)

        lang.language_changed.connect(self.retranslateUi)
        self.retranslateUi()
        self.load_customers()

    def reset_and_search(self):
        self.pagination.set_current_page(1)
        self.load_customers()

    def on_page_changed(self, page: int, page_size: int):
        self.load_customers(page, page_size)

    def get_currency_symbol(self):
        return get_currency_symbol()

    def get_lang(self):
        return lang.get_current()

    # ---------- EXPORT TO EXCEL ----------
    def export_to_excel(self):
        """Export customer list to Excel (.xlsx)"""
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"customer_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Customer List" if self.current_language != "my" else "ဝယ်ယူသူစာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            search_text = self.search_input.text().strip().lower()
            lang_code = self.current_language
            
            conn = connect_db()
            cursor = conn.cursor()
            
            if search_text:
                like = f'%{search_text}%'
                cursor.execute("""
                    SELECT name, phone, email, address, total_visit, total_spent, points,
                           COALESCE(credit_limit, 0), COALESCE(current_balance, 0)
                    FROM customers
                    WHERE name LIKE ? OR phone LIKE ? OR email LIKE ?
                    ORDER BY name
                """, (like, like, like))
            else:
                cursor.execute("""
                    SELECT name, phone, email, address, total_visit, total_spent, points,
                           COALESCE(credit_limit, 0), COALESCE(current_balance, 0)
                    FROM customers
                    ORDER BY name
                """)
            rows = cursor.fetchall()
            
            # Get summary statistics
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_customers,
                    COALESCE(SUM(current_balance), 0) as total_balance,
                    COALESCE(SUM(credit_limit), 0) as total_credit_limit,
                    COALESCE(SUM(total_spent), 0) as total_spent
                FROM customers
            """)
            stats = cursor.fetchone()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Customers"
            
            # Title
            ws.merge_cells('A1:J1')
            ws['A1'] = "CUSTOMER LIST REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            
            if search_text:
                ws['A3'] = f"Search Filter: {search_text}"
                ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Summary row
            ws['A4'] = f"Total Customers: {stats[0] if stats else 0}"
            ws['A4'].font = Font(bold=True, size=11)
            ws['A5'] = f"Total Outstanding Balance: {format_money(stats[1] if stats else 0, symbol)}"
            ws['A5'].font = Font(bold=True, size=11)
            ws['A6'] = f"Total Credit Limit: {format_money(stats[2] if stats else 0, symbol)}"
            ws['A6'].font = Font(bold=True, size=11)
            ws['A7'] = f"Total Spent All Time: {format_money(stats[3] if stats else 0, symbol)}"
            ws['A7'].font = Font(bold=True, size=11)
            
            # Headers
            if lang_code == "my":
                headers = ["အမည်", "ဖုန်း", "အီးမေး", "လိပ်စာ", "အလည်လာရောက်မှု", 
                          "စုစုပေါင်းသုံးစွဲမှု", "အမှတ်", "ခရက်ဒစ်ကန့်သတ်", "လက်ကျန်အကြွေး"]
            else:
                headers = ["Name", "Phone", "Email", "Address", "Total Visit", 
                          "Total Spent", "Points", "Credit Limit", "Current Balance"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            total_visit = 0
            total_spent = 0
            total_points = 0
            total_credit = 0
            total_balance = 0
            
            for row_idx, row_data in enumerate(rows, start=10):
                name, phone, email, address, visit, spent, points, credit, balance = row_data
                
                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=phone or "")
                ws.cell(row=row_idx, column=3, value=email or "")
                ws.cell(row=row_idx, column=4, value=address or "")
                ws.cell(row=row_idx, column=5, value=visit)
                ws.cell(row=row_idx, column=6, value=format_money(spent, symbol))
                ws.cell(row=row_idx, column=7, value=points)
                ws.cell(row=row_idx, column=8, value=format_money(credit, symbol))
                ws.cell(row=row_idx, column=9, value=format_money(balance, symbol))
                
                total_visit += visit
                total_spent += spent
                total_points += points
                total_credit += credit
                total_balance += balance
            
            # Summary row
            summary_row = len(rows) + 11
            ws.cell(row=summary_row, column=4, value="TOTAL").font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=total_visit)
            ws.cell(row=summary_row, column=6, value=format_money(total_spent, symbol))
            ws.cell(row=summary_row, column=7, value=total_points)
            ws.cell(row=summary_row, column=8, value=format_money(total_credit, symbol))
            ws.cell(row=summary_row, column=9, value=format_money(total_balance, symbol))
            
            # Auto adjust columns
            for col in range(1, 10):
                max_length = 0
                for row in range(9, summary_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 40)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def load_customers(self, page=1, page_size=50):
        try:
            symbol = get_currency_symbol()
            search_text = self.search_input.text().strip().lower()
            conn = connect_db()
            cursor = conn.cursor()
            
            if search_text:
                like = f'%{search_text}%'
                cursor.execute("""
                    SELECT COUNT(*) FROM customers 
                    WHERE name LIKE ? OR phone LIKE ? OR email LIKE ?
                """, (like, like, like))
            else:
                cursor.execute("SELECT COUNT(*) FROM customers")
            total_items = cursor.fetchone()[0]
            self.pagination.set_total_items(total_items, emit_signal=False)

            offset = (page - 1) * page_size
            if search_text:
                cursor.execute("""
                    SELECT id, name, phone, email, address, total_visit, total_spent, points, 
                           COALESCE(credit_limit, 0), COALESCE(current_balance, 0)
                    FROM customers
                    WHERE name LIKE ? OR phone LIKE ? OR email LIKE ?
                    ORDER BY name
                    LIMIT ? OFFSET ?
                """, (like, like, like, page_size, offset))
            else:
                cursor.execute("""
                    SELECT id, name, phone, email, address, total_visit, total_spent, points,
                           COALESCE(credit_limit, 0), COALESCE(current_balance, 0)
                    FROM customers
                    ORDER BY name
                    LIMIT ? OFFSET ?
                """, (page_size, offset))
            rows = cursor.fetchall()
            conn.close()
            
            self.table.setRowCount(len(rows))
            for row_idx, row_data in enumerate(rows):
                for col_idx, val in enumerate(row_data):
                    if col_idx == 6:  # total_spent
                        item = QTableWidgetItem(format_money(val, symbol))
                    elif col_idx == 8 or col_idx == 9:  # credit_limit or current_balance
                        item = QTableWidgetItem(format_money(val, symbol))
                        if col_idx == 9 and val > 0:  # current_balance
                            item.setForeground(Qt.GlobalColor.red)
                    else:
                        item = QTableWidgetItem(str(val) if val else "")
                    self.table.setItem(row_idx, col_idx, item)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load customers: {e}")

    def select_customer(self, row, col):
        id_item = self.table.item(row, 0)
        if id_item:
            self.selected_customer_id = int(id_item.text())

    def get_customer_name(self):
        if not self.selected_customer_id:
            return "Unknown"
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM customers WHERE id = ?", (self.selected_customer_id,))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else "Unknown"

    def credit_sale(self):
        if not self.selected_customer_id:
            lang_code = lang.get_current()
            msg = "Please select a customer first." if lang_code != "my" else "ကျေးဇူးပြု၍ ဝယ်ယူသူတစ်ဦးကို ရွေးပါ။"
            QMessageBox.warning(self, "No Selection", msg)
            return
        
        customer_name = self.get_customer_name()
        dialog = CreditSaleDialog(self.selected_customer_id, customer_name, self)
        if dialog.exec():
            self.load_customers()

    def payment_collection(self):
        if not self.selected_customer_id:
            lang_code = lang.get_current()
            msg = "Please select a customer first." if lang_code != "my" else "ကျေးဇူးပြု၍ ဝယ်ယူသူတစ်ဦးကို ရွေးပါ။"
            QMessageBox.warning(self, "No Selection", msg)
            return
        
        customer_name = self.get_customer_name()
        dialog = CreditPaymentDialog(self.selected_customer_id, customer_name, self)
        if dialog.exec():
            self.load_customers()

    def show_ledger(self):
        if not self.selected_customer_id:
            lang_code = lang.get_current()
            msg = "Please select a customer first." if lang_code != "my" else "ကျေးဇူးပြု၍ ဝယ်ယူသူတစ်ဦးကို ရွေးပါ။"
            QMessageBox.warning(self, "No Selection", msg)
            return
        
        customer_name = self.get_customer_name()
        dialog = CustomerLedgerDialog(self.selected_customer_id, customer_name, self)
        dialog.exec()

    def show_outstanding_report(self):
        dialog = OutstandingReportDialog(self)
        dialog.exec()

    def add_customer(self):
        try:
            dialog = AddEditCustomerDialog(language=self.current_language)
            if dialog.exec():
                data = dialog.get_data()
                if not data["name"]:
                    msg = "အမည်ထည့်ရန်လိုအပ်ပါသည်။" if self.current_language == "my" else "Name is required"
                    QMessageBox.warning(self, "Error", msg)
                    return
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO customers (name, phone, email, address, credit_limit, remarks, total_visit, total_spent, points, current_balance)
                    VALUES (?, ?, ?, ?, ?, ?, 0, 0, 0, 0)
                """, (data["name"], data["phone"], data["email"], data["address"], data["credit_limit"], data["remarks"]))
                conn.commit()
                conn.close()
                self.load_customers()
                main_window = self.window()
                if hasattr(main_window, 'current_user'):
                    log_activity(main_window.current_user["id"], main_window.current_user["username"],
                               "Add Customer", f"Customer: {data['name']}")
                msg = "ဝယ်ယူသူထည့်သွင်းပြီးပါပြီ။" if self.current_language == "my" else "Customer added"
                QMessageBox.information(self, "Success", msg)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not add customer: {e}")

    def edit_customer(self):
        if not self.selected_customer_id:
            msg = "ကျေးဇူးပြု၍ ဝယ်ယူသူတစ်ဦးကို ရွေးပါ။" if self.current_language == "my" else "Please select a customer first."
            QMessageBox.warning(self, "No Selection", msg)
            return
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT name, phone, email, address, credit_limit, remarks 
                FROM customers WHERE id = ?
            """, (self.selected_customer_id,))
            row = cursor.fetchone()
            conn.close()
            if not row:
                QMessageBox.warning(self, "Error", "Customer not found")
                return
            old_name = row[0]
            customer_data = {
                "name": old_name,
                "phone": row[1] or "",
                "email": row[2] or "",
                "address": row[3] or "",
                "credit_limit": row[4] or 0,
                "remarks": row[5] or ""
            }
            dialog = AddEditCustomerDialog(customer_data=customer_data, language=self.current_language)
            if dialog.exec():
                data = dialog.get_data()
                if not data["name"]:
                    msg = "အမည်ထည့်ရန်လိုအပ်ပါသည်။" if self.current_language == "my" else "Name is required"
                    QMessageBox.warning(self, "Error", msg)
                    return
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE customers SET name=?, phone=?, email=?, address=?, credit_limit=?, remarks=?
                    WHERE id=?
                """, (data["name"], data["phone"], data["email"], data["address"], data["credit_limit"], data["remarks"], self.selected_customer_id))
                conn.commit()
                conn.close()
                self.load_customers()
                main_window = self.window()
                if hasattr(main_window, 'current_user'):
                    log_activity(main_window.current_user["id"], main_window.current_user["username"],
                               "Edit Customer", f"Customer: {old_name} → {data['name']}")
                msg = "ဝယ်ယူသူပြင်ဆင်ပြီးပါပြီ။" if self.current_language == "my" else "Customer updated"
                QMessageBox.information(self, "Success", msg)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not edit customer: {e}")

    def delete_customer(self):
        if self.user_role not in ['admin']:
            QMessageBox.warning(self, "Access Denied", "You don't have permission to delete customers.")
            return
        if not self.selected_customer_id:
            msg = "ကျေးဇူးပြု၍ ဝယ်ယူသူတစ်ဦးကို ရွေးပါ။" if self.current_language == "my" else "Please select a customer first."
            QMessageBox.warning(self, "No Selection", msg)
            return
        
        # Check if customer has credit sales
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM credit_sales WHERE customer_id = ? AND balance_amount > 0", (self.selected_customer_id,))
        credit_count = cursor.fetchone()[0]
        conn.close()
        
        if credit_count > 0:
            msg = "This customer has outstanding credit sales. Please settle all debts before deleting." if self.current_language != "my" else "ဤဝယ်ယူသူတွင် အကြွေးကျန်ရှိနေပါသည်။ ဖျက်မည်ဆိုပါက အကြွေးအားလုံးကို ဦးစွာရှင်းပါ။"
            QMessageBox.warning(self, "Cannot Delete", msg)
            return
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM customers WHERE id=?", (self.selected_customer_id,))
        cust_name = cursor.fetchone()[0]
        conn.close()
        
        confirm_text = "ဤဝယ်ယူသူကို ဖျက်မည်လား?" if self.current_language == "my" else "Delete this customer?"
        reply = QMessageBox.question(self, "Confirm Delete", confirm_text,
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM customers WHERE id=?", (self.selected_customer_id,))
                conn.commit()
                conn.close()
                self.selected_customer_id = None
                self.load_customers()
                main_window = self.window()
                if hasattr(main_window, 'current_user'):
                    log_activity(main_window.current_user["id"], main_window.current_user["username"],
                               "Delete Customer", f"Customer: {cust_name}")
                msg = "ဝယ်ယူသူဖျက်ပြီးပါပြီ။" if self.current_language == "my" else "Customer deleted"
                QMessageBox.information(self, "Deleted", msg)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not delete customer: {e}")

    def retranslateUi(self):
        lang_code = lang.get_current()
        self.current_language = lang_code
        
        if lang_code == "my":
            self.search_input.setPlaceholderText("အမည်၊ ဖုန်း၊ အီးမေးဖြင့် ရှာရန်...")
            self.btn_add.setText("ဝယ်ယူသူအသစ်")
            self.btn_edit.setText("ပြင်ဆင်မည်")
            self.btn_delete.setText("ဖျက်မည်")
            self.btn_credit_sale.setText("အကြွေးရောင်းမည်")
            self.btn_payment.setText("ငွေပေးချေမှုကောက်ခံမည်")
            self.btn_ledger.setText("စာရင်း")
            self.btn_outstanding_report.setText("အကြွေးကျန်စာရင်း")
            self.btn_export_excel.setText("📊 Excel ထုတ်မည်")
            headers = [
                "ID", "အမည်", "ဖုန်း", "အီးမေး", "လိပ်စာ",
                "အလည်လာရောက်မှု", "စုစုပေါင်းသုံးစွဲမှု", 
                "အမှတ်", "ခရက်ဒစ်ကန့်သတ်", "လက်ကျန်အကြွေး"
            ]
        else:
            self.search_input.setPlaceholderText("Search by name, phone, email...")
            self.btn_add.setText("Add Customer")
            self.btn_edit.setText("Edit")
            self.btn_delete.setText("Delete")
            self.btn_credit_sale.setText("Credit Sale")
            self.btn_payment.setText("Payment Collection")
            self.btn_ledger.setText("Ledger")
            self.btn_outstanding_report.setText("Outstanding Report")
            self.btn_export_excel.setText("📊 Export Excel")
            headers = [
                "ID", "Name", "Phone", "Email", "Address",
                "Total Visit", "Total Spent", "Points", "Credit Limit", "Current Balance"
            ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

    def showEvent(self, event):
        self.load_customers()
        super().showEvent(event)