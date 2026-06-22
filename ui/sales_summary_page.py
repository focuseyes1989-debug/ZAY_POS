# sales_summary_page.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget,
    QTableWidgetItem, QHeaderView, QDateEdit, QPushButton, QTabWidget,
    QSizePolicy, QFileDialog, QMessageBox, QLineEdit
)
from PyQt6.QtCore import Qt, QDate
from models.database import connect_db
from utils.currency import format_money, get_currency_symbol
from utils.language import lang
from utils.excel_exporter import ExcelExporter
from datetime import datetime
import csv


class SalesSummaryPage(QWidget):
    def __init__(self):
        super().__init__()
        main_layout = QVBoxLayout()

        # Combined row: quick buttons + date pickers + refresh + export
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.btn_today = QPushButton("Today")
        self.btn_this_week = QPushButton("This Week")
        self.btn_this_month = QPushButton("This Month")
        self.btn_last_month = QPushButton("Last Month")
        self.btn_this_year = QPushButton("This Year")

        self.btn_today.clicked.connect(self.set_today_range)
        self.btn_this_week.clicked.connect(self.set_this_week_range)
        self.btn_this_month.clicked.connect(self.set_this_month_range)
        self.btn_last_month.clicked.connect(self.set_last_month_range)
        self.btn_this_year.clicked.connect(self.set_this_year_range)

        filter_layout.addWidget(self.btn_today)
        filter_layout.addWidget(self.btn_this_week)
        filter_layout.addWidget(self.btn_this_month)
        filter_layout.addWidget(self.btn_last_month)
        filter_layout.addWidget(self.btn_this_year)
        filter_layout.addStretch()

        self.from_label = QLabel("From:")
        filter_layout.addWidget(self.from_label)
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-30))
        self.from_date.dateChanged.connect(self.load_all_tabs)
        filter_layout.addWidget(self.from_date)

        self.to_label = QLabel("To:")
        filter_layout.addWidget(self.to_label)
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        self.to_date.dateChanged.connect(self.load_all_tabs)
        filter_layout.addWidget(self.to_date)

        self.btn_refresh = QPushButton("Refresh")
        self.btn_refresh.clicked.connect(self.load_all_tabs)
        filter_layout.addWidget(self.btn_refresh)

        # Export buttons
        self.btn_export_excel = QPushButton("📊 Export Excel")
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        filter_layout.addWidget(self.btn_export_excel)

        filter_layout.addStretch()
        main_layout.addLayout(filter_layout)

        # Tabs
        self.tabs = QTabWidget()
        self.tab_names = {
            0: "Top 20 Sales by Item",
            1: "Sales by Item",
            2: "Sales by Category",
            3: "Sales by Payment Type"
        }
        self.tabs.addTab(self.create_top_items_tab(), self.tab_names[0])
        self.tabs.addTab(self.create_items_tab(), self.tab_names[1])
        self.tabs.addTab(self.create_categories_tab(), self.tab_names[2])
        self.tabs.addTab(self.create_payment_tab(), self.tab_names[3])

        main_layout.addWidget(self.tabs)
        self.setLayout(main_layout)

        # Store full data for searching
        self.items_data = []
        self.categories_data = []

        lang.language_changed.connect(self.retranslateUi)
        self.load_all_tabs()
        self.retranslateUi()

    def showEvent(self, event):
        self.load_all_tabs()
        super().showEvent(event)

    # ---------- Quick date range helpers ----------
    def set_today_range(self):
        today = QDate.currentDate()
        self.from_date.setDate(today)
        self.to_date.setDate(today)
        self.load_all_tabs()

    def set_this_week_range(self):
        today = QDate.currentDate()
        start_of_week = today.addDays(-(today.dayOfWeek() - 1))
        end_of_week = start_of_week.addDays(6)
        self.from_date.setDate(start_of_week)
        self.to_date.setDate(end_of_week)
        self.load_all_tabs()

    def set_this_month_range(self):
        today = QDate.currentDate()
        start_of_month = QDate(today.year(), today.month(), 1)
        end_of_month = QDate(today.year(), today.month(), today.daysInMonth())
        self.from_date.setDate(start_of_month)
        self.to_date.setDate(end_of_month)
        self.load_all_tabs()

    def set_last_month_range(self):
        today = QDate.currentDate()
        first_day_this_month = QDate(today.year(), today.month(), 1)
        last_day_last_month = first_day_this_month.addDays(-1)
        first_day_last_month = QDate(last_day_last_month.year(), last_day_last_month.month(), 1)
        self.from_date.setDate(first_day_last_month)
        self.to_date.setDate(last_day_last_month)
        self.load_all_tabs()

    def set_this_year_range(self):
        today = QDate.currentDate()
        start_of_year = QDate(today.year(), 1, 1)
        end_of_year = QDate(today.year(), 12, 31)
        self.from_date.setDate(start_of_year)
        self.to_date.setDate(end_of_year)
        self.load_all_tabs()

    def get_date_range(self):
        from_date = self.from_date.date().toString("yyyy-MM-dd")
        to_date = self.to_date.date().toString("yyyy-MM-dd")
        return from_date, to_date

    def get_currency_symbol(self):
        return get_currency_symbol()

    def get_lang(self):
        return lang.get_current()

    # ---------- EXPORT TO EXCEL ----------
    def export_to_excel(self):
        """Export all sales summary tabs to Excel (.xlsx)"""
        from_date, to_date = self.get_date_range()
        symbol = self.get_currency_symbol()
        lang_code = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"sales_summary_{from_date}_to_{to_date}.xlsx",
            "Export Sales Summary" if lang_code != "my" else "ရောင်းအားအကျဉ်းချုပ် ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Tab 1: Top Products
            ws1 = wb.active
            ws1.title = "Top Products"
            self._export_top_products_to_worksheet(ws1, from_date, to_date, symbol, lang_code)
            
            # Tab 2: Products Detail
            ws2 = wb.create_sheet("Products Detail")
            self._export_products_detail_to_worksheet(ws2, from_date, to_date, symbol, lang_code)
            
            # Tab 3: Categories
            ws3 = wb.create_sheet("Categories")
            self._export_categories_to_worksheet(ws3, from_date, to_date, symbol, lang_code)
            
            # Tab 4: Payment Types
            ws4 = wb.create_sheet("Payment Types")
            self._export_payment_to_worksheet(ws4, from_date, to_date, symbol, lang_code)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def _export_top_products_to_worksheet(self, ws, from_date, to_date, symbol, lang_code):
        """Export top 20 products to worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sale_items.product_name, COALESCE(SUM(sale_items.total), 0) as total_sales
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY sale_items.product_name
            ORDER BY total_sales DESC
            LIMIT 20
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = "TOP 20 SALES BY ITEM"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A2'].font = Font(size=10, color="7f8c8d")
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(size=10, color="7f8c8d")
        
        # Headers
        headers = ["ပစ္စည်းအမည်", "စုစုပေါင်းရောင်းအား"] if lang_code == "my" else ["Product Name", "Total Sales"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_idx, (name, total) in enumerate(rows, start=6):
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=format_money(total, symbol))
        
        # Adjust columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _export_products_detail_to_worksheet(self, ws, from_date, to_date, symbol, lang_code):
        """Export products detail to worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sale_items.product_name,
                   COALESCE(products.category, 'Uncategorized') as category,
                   COALESCE(SUM(sale_items.qty), 0) as total_qty,
                   COALESCE(SUM(sale_items.total), 0) as net_sales,
                   COALESCE(SUM(products.cost * sale_items.qty), 0) as cogs,
                   COALESCE(SUM(sale_items.total) - SUM(products.cost * sale_items.qty), 0) as gross_profit
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            LEFT JOIN products ON sale_items.product_name = products.name
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY sale_items.product_name
            ORDER BY sale_items.product_name
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "SALES BY PRODUCT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Headers
        if lang_code == "my":
            headers = ["ပစ္စည်းအမည်", "အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", 
                      "အသားတင်ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Product Name", "Category", "Items Sold", "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_qty = 0
        total_sales = 0
        total_cogs = 0
        total_profit = 0
        
        for row_idx, row_data in enumerate(rows, start=6):
            name, category, qty, sales, cogs, profit = row_data
            
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=category)
            ws.cell(row=row_idx, column=3, value=qty)
            ws.cell(row=row_idx, column=4, value=format_money(sales, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=6, value=format_money(profit, symbol))
            
            total_qty += qty
            total_sales += sales
            total_cogs += cogs
            total_profit += profit
        
        # Summary row
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=total_qty)
        ws.cell(row=summary_row, column=4, value=format_money(total_sales, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=6, value=format_money(total_profit, symbol))
        
        # Adjust columns
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_categories_to_worksheet(self, ws, from_date, to_date, symbol, lang_code):
        """Export categories summary to worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(products.category, 'Uncategorized') as category,
                   COALESCE(SUM(sale_items.qty), 0) as items_sold,
                   COALESCE(SUM(sale_items.total), 0) as net_sales,
                   COALESCE(SUM(products.cost * sale_items.qty), 0) as cogs
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            LEFT JOIN products ON sale_items.product_name = products.name
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY products.category
            ORDER BY net_sales DESC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "SALES BY CATEGORY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Headers
        if lang_code == "my":
            headers = ["အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", "စုစုပေါင်းရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Category", "Items Sold", "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_items = 0
        total_sales = 0
        total_cogs = 0
        total_profit = 0
        
        for row_idx, row_data in enumerate(rows, start=6):
            category, items, sales, cogs = row_data
            profit = sales - cogs
            
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=items)
            ws.cell(row=row_idx, column=3, value=format_money(sales, symbol))
            ws.cell(row=row_idx, column=4, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(profit, symbol))
            
            total_items += items
            total_sales += sales
            total_cogs += cogs
            total_profit += profit
        
        # Summary row
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_items)
        ws.cell(row=summary_row, column=3, value=format_money(total_sales, symbol))
        ws.cell(row=summary_row, column=4, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_profit, symbol))
        
        # Adjust columns
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_payment_to_worksheet(self, ws, from_date, to_date, symbol, lang_code):
        """Export payment types summary to worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(payment_type, 'Other') as payment_type,
                   COUNT(*) as transaction_count,
                   COALESCE(SUM(total), 0) as total_amount
            FROM sales
            WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            GROUP BY payment_type
            ORDER BY payment_type
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = "SALES BY PAYMENT TYPE"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Headers
        if lang_code == "my":
            headers = ["ငွေပေးချေမှုအမျိုးအစား", "ငွေပေးချေမှုအရေအတွက်", "ငွေပေးချေမှုပမာဏ"]
        else:
            headers = ["Payment Type", "Transaction Count", "Amount"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_count = 0
        total_amount = 0
        
        for row_idx, row_data in enumerate(rows, start=6):
            ptype, count, amount = row_data
            
            ws.cell(row=row_idx, column=1, value=ptype)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=format_money(amount, symbol))
            
            total_count += count
            total_amount += amount
        
        # Summary row
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_count)
        ws.cell(row=summary_row, column=3, value=format_money(total_amount, symbol))
        
        # Adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20

    # ---------- Tab 1: Top 20 Sales by Item ----------
    def create_top_items_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        self.top_items_table = QTableWidget()
        self.top_items_table.setColumnCount(2)
        self.top_items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self.top_items_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.top_items_table)
        widget.setLayout(layout)
        return widget

    def load_top_items_tab(self):
        from_date, to_date = self.get_date_range()
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sale_items.product_name, COALESCE(SUM(sale_items.total), 0) as total_sales
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY sale_items.product_name
            ORDER BY total_sales DESC
            LIMIT 20
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        self.top_items_table.setRowCount(0)
        for name, total in rows:
            r = self.top_items_table.rowCount()
            self.top_items_table.insertRow(r)
            self.top_items_table.setItem(r, 0, QTableWidgetItem(name))
            self.top_items_table.setItem(r, 1, QTableWidgetItem(format_money(total)))
        if self.get_lang() == "my":
            self.top_items_table.setHorizontalHeaderLabels(["ပစ္စည်းအမည်", "စုစုပေါင်းရောင်းအား"])
        else:
            self.top_items_table.setHorizontalHeaderLabels(["Product Name", "Total Sales"])

    # ---------- Tab 2: Sales by Item (with Search) ----------
    def create_items_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Search bar
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 Search:"))
        self.items_search = QLineEdit()
        self.items_search.setPlaceholderText("Search product name...")
        self.items_search.textChanged.connect(self.filter_items_table)
        self.items_search.setStyleSheet("""
            QLineEdit {
                padding: 6px 10px;
                border: 1px solid #ccc;
                border-radius: 4px;
                min-width: 200px;
            }
            QLineEdit:focus {
                border-color: #5865f2;
            }
        """)
        search_layout.addWidget(self.items_search)
        search_layout.addStretch()
        
        # Clear button
        self.items_clear_btn = QPushButton("✕ Clear")
        self.items_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.items_clear_btn.clicked.connect(self.clear_items_search)
        search_layout.addWidget(self.items_clear_btn)
        
        layout.addLayout(search_layout)
        
        # Table
        self.items_table = QTableWidget()
        self.items_table.setColumnCount(6)
        self.items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self.items_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.items_table)
        
        widget.setLayout(layout)
        return widget

    def clear_items_search(self):
        self.items_search.clear()

    def filter_items_table(self):
        """Filter items table based on search text"""
        search_text = self.items_search.text().lower().strip()
        
        if not search_text:
            # Show all rows
            self.load_items_tab()
            return
        
        # Filter the stored data
        filtered_rows = []
        for row_data in self.items_data:
            name = row_data[0].lower()
            if search_text in name:
                filtered_rows.append(row_data)
        
        self._display_filtered_items(filtered_rows)

    def _display_filtered_items(self, rows):
        """Display filtered items in the table"""
        symbol = get_currency_symbol()
        self.items_table.setRowCount(0)
        
        for row_data in rows:
            r = self.items_table.rowCount()
            self.items_table.insertRow(r)
            self.items_table.setItem(r, 0, QTableWidgetItem(row_data[0]))
            self.items_table.setItem(r, 1, QTableWidgetItem(row_data[1]))
            self.items_table.setItem(r, 2, QTableWidgetItem(str(row_data[2])))
            self.items_table.setItem(r, 3, QTableWidgetItem(format_money(row_data[3], symbol)))
            self.items_table.setItem(r, 4, QTableWidgetItem(format_money(row_data[4], symbol)))
            self.items_table.setItem(r, 5, QTableWidgetItem(format_money(row_data[5], symbol)))
        
        # Update headers if needed
        if self.get_lang() == "my":
            self.items_table.setHorizontalHeaderLabels([
                "ပစ္စည်းအမည်", "အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်",
                "အသားတင်ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"
            ])
        else:
            self.items_table.setHorizontalHeaderLabels([
                "Product Name", "Category", "Items Sold", "Net Sales", "Cost of Goods", "Gross Profit"
            ])

    def load_items_tab(self):
        from_date, to_date = self.get_date_range()
        symbol = get_currency_symbol()
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT sale_items.product_name,
                   COALESCE(products.category, 'Uncategorized') as category,
                   COALESCE(SUM(sale_items.qty), 0) as total_qty,
                   COALESCE(SUM(sale_items.total), 0) as net_sales,
                   COALESCE(SUM(products.cost * sale_items.qty), 0) as cogs,
                   COALESCE(SUM(sale_items.total) - SUM(products.cost * sale_items.qty), 0) as gross_profit
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            LEFT JOIN products ON sale_items.product_name = products.name
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY sale_items.product_name
            ORDER BY sale_items.product_name
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Store data for searching
        self.items_data = [list(row) for row in rows]
        
        # Check if search is active
        if self.items_search.text().strip():
            self.filter_items_table()
        else:
            self._display_filtered_items(self.items_data)

    # ---------- Tab 3: Sales by Category (with Search) ----------
    def create_categories_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Search bar
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 Search:"))
        self.categories_search = QLineEdit()
        self.categories_search.setPlaceholderText("Search category name...")
        self.categories_search.textChanged.connect(self.filter_categories_table)
        self.categories_search.setStyleSheet("""
            QLineEdit {
                padding: 6px 10px;
                border: 1px solid #ccc;
                border-radius: 4px;
                min-width: 200px;
            }
            QLineEdit:focus {
                border-color: #5865f2;
            }
        """)
        search_layout.addWidget(self.categories_search)
        search_layout.addStretch()
        
        # Clear button
        self.categories_clear_btn = QPushButton("✕ Clear")
        self.categories_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.categories_clear_btn.clicked.connect(self.clear_categories_search)
        search_layout.addWidget(self.categories_clear_btn)
        
        layout.addLayout(search_layout)
        
        # Table
        self.categories_table = QTableWidget()
        self.categories_table.setColumnCount(5)
        self.categories_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self.categories_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.categories_table)
        
        widget.setLayout(layout)
        return widget

    def clear_categories_search(self):
        self.categories_search.clear()

    def filter_categories_table(self):
        """Filter categories table based on search text"""
        search_text = self.categories_search.text().lower().strip()
        
        if not search_text:
            # Show all rows
            self.load_categories_tab()
            return
        
        # Filter the stored data
        filtered_rows = []
        for row_data in self.categories_data:
            category = row_data[0].lower()
            if search_text in category:
                filtered_rows.append(row_data)
        
        self._display_filtered_categories(filtered_rows)

    def _display_filtered_categories(self, rows):
        """Display filtered categories in the table"""
        symbol = get_currency_symbol()
        self.categories_table.setRowCount(0)
        
        for row_data in rows:
            r = self.categories_table.rowCount()
            self.categories_table.insertRow(r)
            self.categories_table.setItem(r, 0, QTableWidgetItem(row_data[0] if row_data[0] else "Uncategorized"))
            self.categories_table.setItem(r, 1, QTableWidgetItem(str(row_data[1])))
            self.categories_table.setItem(r, 2, QTableWidgetItem(format_money(row_data[2], symbol)))
            self.categories_table.setItem(r, 3, QTableWidgetItem(format_money(row_data[3], symbol)))
            profit = row_data[2] - row_data[3]
            self.categories_table.setItem(r, 4, QTableWidgetItem(format_money(profit, symbol)))
        
        # Update headers if needed
        if self.get_lang() == "my":
            self.categories_table.setHorizontalHeaderLabels([
                "အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", "စုစုပေါင်းရောင်းအား",
                "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"
            ])
        else:
            self.categories_table.setHorizontalHeaderLabels([
                "Category", "Items Sold", "Net Sales", "Cost of Goods", "Gross Profit"
            ])

    def load_categories_tab(self):
        from_date, to_date = self.get_date_range()
        symbol = get_currency_symbol()
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(products.category, 'Uncategorized') as category,
                   COALESCE(SUM(sale_items.qty), 0) as items_sold,
                   COALESCE(SUM(sale_items.total), 0) as net_sales,
                   COALESCE(SUM(products.cost * sale_items.qty), 0) as cogs
            FROM sale_items
            JOIN sales ON sale_items.sale_id = sales.id
            LEFT JOIN products ON sale_items.product_name = products.name
            WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            GROUP BY products.category
            ORDER BY net_sales DESC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        # Store data for searching
        self.categories_data = [list(row) for row in rows]
        
        # Check if search is active
        if self.categories_search.text().strip():
            self.filter_categories_table()
        else:
            self._display_filtered_categories(self.categories_data)

    # ---------- Tab 4: Sales by Payment Type ----------
    def create_payment_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        self.payment_table = QTableWidget()
        self.payment_table.setColumnCount(3)
        self.payment_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self.payment_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.payment_table)
        widget.setLayout(layout)
        return widget

    def load_payment_tab(self):
        from_date, to_date = self.get_date_range()
        symbol = get_currency_symbol()
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(payment_type, 'Other') as payment_type,
                   COUNT(*) as transaction_count,
                   COALESCE(SUM(total), 0) as total_amount
            FROM sales
            WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            GROUP BY payment_type
            ORDER BY payment_type
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        self.payment_table.setRowCount(0)
        total_count = 0
        total_amount = 0.0
        for ptype, count, amount in rows:
            r = self.payment_table.rowCount()
            self.payment_table.insertRow(r)
            self.payment_table.setItem(r, 0, QTableWidgetItem(ptype))
            self.payment_table.setItem(r, 1, QTableWidgetItem(str(count)))
            self.payment_table.setItem(r, 2, QTableWidgetItem(format_money(amount, symbol)))
            total_count += count
            total_amount += amount
        # Total row
        r = self.payment_table.rowCount()
        self.payment_table.insertRow(r)
        font = self.payment_table.font()
        font.setBold(True)
        total_item = QTableWidgetItem("TOTAL")
        total_item.setFont(font)
        self.payment_table.setItem(r, 0, total_item)
        item = QTableWidgetItem(str(total_count))
        item.setFont(font)
        self.payment_table.setItem(r, 1, item)
        item = QTableWidgetItem(format_money(total_amount, symbol))
        item.setFont(font)
        self.payment_table.setItem(r, 2, item)
        
        if self.get_lang() == "my":
            self.payment_table.setHorizontalHeaderLabels([
                "ငွေပေးချေမှုအမျိုးအစား", "ငွေပေးချေမှုအရေအတွက်", "ငွေပေးချေမှုပမာဏ"
            ])
        else:
            self.payment_table.setHorizontalHeaderLabels([
                "Payment Type", "Transaction Count", "Amount"
            ])

    def load_all_tabs(self):
        self.load_top_items_tab()
        self.load_items_tab()
        self.load_categories_tab()
        self.load_payment_tab()

    def retranslateUi(self):
        lang_code = self.get_lang()
        if lang_code == "my":
            self.btn_today.setText("ယနေ့")
            self.btn_this_week.setText("ဤတစ်ပတ်")
            self.btn_this_month.setText("ဤလ")
            self.btn_last_month.setText("ပြီးခဲ့သည့်လ")
            self.btn_this_year.setText("ဤနှစ်")
            self.from_label.setText("မှ:")
            self.to_label.setText("အထိ:")
            self.btn_refresh.setText("ပြန်လည်")
            self.btn_export_excel.setText("📊 Excel ထုတ်မည်")
            
            # Update search placeholders
            self.items_search.setPlaceholderText("ပစ္စည်းအမည်ရှာရန်...")
            self.categories_search.setPlaceholderText("အမျိုးအစားရှာရန်...")
            self.items_clear_btn.setText("✕ ရှင်းမည်")
            self.categories_clear_btn.setText("✕ ရှင်းမည်")

            tab_titles_my = {
                0: "ထိပ်ဆုံးရောင်းအားရှိပစ္စည်း ၂၀",
                1: "ပစ္စည်းအလိုက်ရောင်းအား",
                2: "အမျိုးအစားအလိုက်ရောင်းအား",
                3: "ငွေပေးချေမှုအလိုက်ရောင်းအား"
            }
            for idx, title in tab_titles_my.items():
                self.tabs.setTabText(idx, title)
        else:
            self.btn_today.setText("Today")
            self.btn_this_week.setText("This Week")
            self.btn_this_month.setText("This Month")
            self.btn_last_month.setText("Last Month")
            self.btn_this_year.setText("This Year")
            self.from_label.setText("From:")
            self.to_label.setText("To:")
            self.btn_refresh.setText("Refresh")
            self.btn_export_excel.setText("📊 Export Excel")
            
            # Update search placeholders
            self.items_search.setPlaceholderText("Search product name...")
            self.categories_search.setPlaceholderText("Search category name...")
            self.items_clear_btn.setText("✕ Clear")
            self.categories_clear_btn.setText("✕ Clear")

            for idx, title in self.tab_names.items():
                self.tabs.setTabText(idx, title)

        self.load_top_items_tab()
        self.load_items_tab()
        self.load_categories_tab()
        self.load_payment_tab()