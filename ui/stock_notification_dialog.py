# stock_notification_dialog.py
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, 
    QHeaderView, QPushButton, QLabel, QApplication, 
    QHBoxLayout, QMessageBox, QLineEdit, QComboBox
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFontMetrics, QIcon, QKeySequence
from models.database import connect_db
from utils.language import lang
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from ui.widgets.pagination_widget import PaginationWidget
from ui.product_detail_dialog import ProductDetailDialog  # ✅ Import ProductDetailDialog
from datetime import datetime
from loguru import logger


class StockNotificationDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Stock Alerts")
        self.setMinimumSize(900, 550)
        self.setModal(False)
        
        # Pagination variables
        self.current_page = 1
        self.page_size = 50
        self.total_items = 0
        self.current_rows = []
        self.all_rows = []  # Store all filtered rows
        
        # ✅ Store product IDs for double click
        self.product_ids = []

        layout = QVBoxLayout()
        
        # Top section with title and export button
        top_layout = QHBoxLayout()
        self.label = QLabel()
        top_layout.addWidget(self.label)
        top_layout.addStretch()
        
        self.btn_export = QPushButton("📊 Export Excel")
        self.btn_export.clicked.connect(self.export_to_excel)
        top_layout.addWidget(self.btn_export)
        
        layout.addLayout(top_layout)

        # Filter section
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name, SKU or barcode...")
        self.search_input.textChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_input, 2)
        
        # Status filter
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All Status", "Out of Stock", "Low Stock"])
        self.status_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(QLabel("Status:"))
        filter_layout.addWidget(self.status_filter, 1)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(7)  # ✅ Added ID column (hidden)
        self.table.setHorizontalHeaderLabels(["ID", "Product", "SKU", "Current Stock", "Low Stock Level", "Status", "Suggested Order"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        # ✅ Hide ID column
        self.table.setColumnHidden(0, True)
        
        # Set column stretch modes for responsiveness
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # ID (hidden)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Product
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # SKU
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Stock
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Low Stock
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Status
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Suggested
        
        # Enable word wrap for product name column
        self.table.setWordWrap(True)
        
        # Set row height to adjust for wrapped text
        self.table.verticalHeader().setDefaultSectionSize(40)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        
        # ✅ Connect double click signal
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        
        layout.addWidget(self.table)

        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        
        # Load data
        self.load_all_data()
        self.apply_filters_and_pagination()
        
        # Connect resize event to adjust column width
        self.table.horizontalHeader().sectionResized.connect(self.on_column_resized)
        
        # Use timer to handle initial resize
        QTimer.singleShot(100, self.adjust_product_column)
        
        # Language support
        lang.language_changed.connect(self.retranslateUi)
        self.retranslateUi()
        
        # Set window flags to show close button in title bar only
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint
        )

    def retranslateUi(self):
        """Update UI text when language changes"""
        lang_code = lang.get_current()
        if lang_code == "my":
            self.btn_export.setText("📊 Excel ထုတ်မည်")
            self.btn_export.setToolTip("စတော့သတိပေးချက်စာရင်းကို Excel ဖိုင်အဖြစ် ထုတ်ယူမည်")
            self.search_input.setPlaceholderText("ပစ္စည်းအမည် / SKU / ဘားကုဒ်ဖြင့် ရှာရန်...")
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "ကုန်သွားပြီ")
            self.status_filter.setItemText(2, "စတော့နည်းနေပြီ")
        else:
            self.btn_export.setText("📊 Export Excel")
            self.btn_export.setToolTip("Export stock alerts to Excel file")
            self.search_input.setPlaceholderText("Search by name, SKU or barcode...")
            self.status_filter.setItemText(0, "All Status")
            self.status_filter.setItemText(1, "Out of Stock")
            self.status_filter.setItemText(2, "Low Stock")
        
        # Update pagination text
        self.pagination._update_controls()
        
        # Update table headers
        self.update_table_headers()

    def update_table_headers(self):
        """Update table headers based on language"""
        lang_code = lang.get_current()
        if lang_code == "my":
            self.table.setHorizontalHeaderLabels([
                "ID", "ပစ္စည်းအမည်", "SKU", "လက်ကျန်", 
                "သတိပေးပမာဏ", "အခြေအနေ", "ပြန်မှာသင့်ပမာဏ"
            ])
        else:
            self.table.setHorizontalHeaderLabels([
                "ID", "Product", "SKU", "Current Stock", 
                "Low Stock Level", "Status", "Suggested Order"
            ])
        
        # ✅ Keep ID column hidden
        self.table.setColumnHidden(0, True)

    def on_filter_changed(self):
        """Handle filter changes"""
        self.current_page = 1
        self.apply_filters_and_pagination()

    def on_page_changed(self, page: int, page_size: int):
        """Handle page changes from pagination widget"""
        self.current_page = page
        self.page_size = page_size
        self.apply_filters_and_pagination()

    def load_all_data(self):
        """Load all stock alert data from database"""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                id,
                name, 
                sku, 
                stock, 
                low_stock, 
                sold_by,
                CASE 
                    WHEN COALESCE(stock, 0) = 0 THEN 'Out of Stock'
                    WHEN COALESCE(stock, 0) <= COALESCE(low_stock, 0) THEN 'Low Stock'
                    ELSE 'OK'
                END as status,
                (COALESCE(low_stock, 0) * 2) as suggested,
                price,
                category,
                barcode
            FROM products
            WHERE (sold_by IS NULL OR sold_by != 'Service')
              AND (COALESCE(stock, 0) = 0 OR COALESCE(stock, 0) <= COALESCE(low_stock, 0))
            ORDER BY 
                CASE 
                    WHEN COALESCE(stock, 0) = 0 THEN 0
                    ELSE 1
                END,
                name ASC
        """)
        self.all_rows = cursor.fetchall()
        conn.close()

    def apply_filters_and_pagination(self):
        """Apply search filter, status filter and pagination"""
        search_text = self.search_input.text().strip().lower()
        status_filter = self.status_filter.currentText()
        
        # Filter rows
        filtered_rows = []
        self.product_ids = []  # ✅ Reset product IDs
        
        for row in self.all_rows:
            product_id, name, sku, stock, low_stock, sold_by, status, suggested, price, category, barcode = row
            
            # Search filter
            if search_text:
                search_match = (
                    search_text in name.lower() or 
                    (sku and search_text in sku.lower()) or
                    (barcode and search_text in barcode.lower())
                )
                if not search_match:
                    continue
            
            # Status filter
            if status_filter == "Out of Stock" or status_filter == "ကုန်သွားပြီ":
                if status != "Out of Stock":
                    continue
            elif status_filter == "Low Stock" or status_filter == "စတော့နည်းနေပြီ":
                if status != "Low Stock":
                    continue
            
            filtered_rows.append(row)
            self.product_ids.append(product_id)  # ✅ Store product ID
        
        # Update total items
        self.total_items = len(filtered_rows)
        self.pagination.set_total_items(self.total_items, emit_signal=False)
        
        # Calculate pagination
        start_idx = (self.current_page - 1) * self.page_size
        end_idx = min(start_idx + self.page_size, self.total_items)
        page_rows = filtered_rows[start_idx:end_idx]
        page_product_ids = self.product_ids[start_idx:end_idx]  # ✅ Get IDs for current page
        
        # Store current rows for export
        self.current_rows = filtered_rows
        
        # Populate table
        self.populate_table(page_rows, page_product_ids, len(filtered_rows))

    def populate_table(self, rows, product_ids, total_count):
        """Populate table with filtered and paginated data"""
        self.table.setRowCount(len(rows))
        
        lang_code = lang.get_current()
        
        # Update label
        if lang_code == "my":
            self.label.setText(f"စတော့သတိပေးချက် - ပစ္စည်း {total_count} မျိုး")
        else:
            self.label.setText(f"Stock Alerts – {total_count} product(s)")
        
        for row_idx, row in enumerate(rows):
            product_id, name, sku, stock, low_stock, sold_by, status, suggested, price, category, barcode = row
            
            # ✅ Column 0: ID (hidden)
            id_item = QTableWidgetItem(str(product_id))
            id_item.setFlags(id_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 0, id_item)
            
            # Column 1: Product name with word wrap
            name_item = QTableWidgetItem(name)
            name_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 1, name_item)
            
            # Column 2: SKU
            sku_item = QTableWidgetItem(sku or "")
            sku_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 2, sku_item)
            
            # Column 3: Stock
            stock_item = QTableWidgetItem(str(stock))
            stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 3, stock_item)
            
            # Column 4: Low Stock Level
            low_item = QTableWidgetItem(str(low_stock))
            low_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 4, low_item)
            
            # Column 5: Status
            status_display = status
            if lang_code == "my":
                if status == "Out of Stock":
                    status_display = "ကုန်သွားပြီ"
                elif status == "Low Stock":
                    status_display = "စတော့နည်းနေပြီ"
            
            status_item = QTableWidgetItem(status_display)
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            
            # Color code status
            if status == "Out of Stock":
                status_item.setBackground(Qt.GlobalColor.red)
                status_item.setForeground(Qt.GlobalColor.white)
            elif status == "Low Stock":
                status_item.setBackground(Qt.GlobalColor.yellow)
                status_item.setForeground(Qt.GlobalColor.black)
            
            self.table.setItem(row_idx, 5, status_item)
            
            # Column 6: Suggested Order
            suggested_item = QTableWidgetItem(str(suggested))
            suggested_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, 6, suggested_item)
        
        # Update table headers
        self.update_table_headers()
        
        # Apply alternating row colors
        self.table.setAlternatingRowColors(True)
        
        # Adjust product column after loading data
        QTimer.singleShot(50, self.adjust_product_column)

    # ============================================================
    # ✅ DOUBLE CLICK HANDLER - Open Product Detail Dialog
    # ============================================================
    def on_cell_double_clicked(self, row, column):
        """
        Handle double click on table row.
        Opens ProductDetailDialog for the selected product.
        """
        try:
            # Get product ID from the hidden column
            id_item = self.table.item(row, 0)
            if not id_item:
                return
            
            product_id = int(id_item.text())
            
            # Open ProductDetailDialog
            dialog = ProductDetailDialog(product_id)
            dialog.exec()
            
        except ValueError as e:
            logger.error(f"Invalid product ID: {e}")
            QMessageBox.warning(self, "Error", "Could not open product details.")
        except Exception as e:
            logger.error(f"Error opening product detail: {e}")
            QMessageBox.warning(self, "Error", f"Failed to open product details: {str(e)}")

    def adjust_product_column(self):
        """Adjust product column to show full product names"""
        header = self.table.horizontalHeader()
        
        # Get available width for product column
        total_width = self.table.viewport().width()
        fixed_width = 0
        
        # Calculate fixed columns width (skip ID column which is hidden)
        for col in range(1, self.table.columnCount()):
            if header.sectionResizeMode(col) == QHeaderView.ResizeMode.ResizeToContents:
                fixed_width += header.sectionSize(col)
            else:
                fixed_width += header.sectionSize(col)
        
        # Set product column to take remaining space
        product_width = total_width - fixed_width - 20  # Subtract padding
        if product_width > 100:  # Minimum width
            header.resizeSection(1, product_width)
        
        # Update row heights for wrapped text
        self.adjust_row_heights()

    def adjust_row_heights(self):
        """Adjust row heights based on content"""
        font_metrics = QFontMetrics(self.table.font())
        product_col_width = self.table.horizontalHeader().sectionSize(1) - 10  # Subtract padding
        
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 1)  # Product column
            if item:
                text = item.text()
                # Calculate required height for wrapped text
                lines = font_metrics.boundingRect(0, 0, product_col_width, 0, 
                                                   Qt.TextFlag.TextWordWrap, text)
                required_height = lines.height() + 20  # Add padding
                current_height = self.table.verticalHeader().sectionSize(row)
                if required_height > current_height:
                    self.table.verticalHeader().resizeSection(row, required_height)
                elif required_height < current_height and required_height > 20:
                    self.table.verticalHeader().resizeSection(row, required_height)

    def on_column_resized(self, index, old_size, new_size):
        """Handle column resize events"""
        if index == 1:  # Product column
            self.adjust_row_heights()

    def resizeEvent(self, event):
        """Handle dialog resize event"""
        super().resizeEvent(event)
        self.adjust_product_column()

    def showEvent(self, event):
        """Handle dialog show event"""
        super().showEvent(event)
        QTimer.singleShot(50, self.adjust_product_column)
        
        # Reload data when dialog shows
        self.load_all_data()
        self.apply_filters_and_pagination()

    def export_to_excel(self):
        """Export stock alerts to Excel file"""
        if not hasattr(self, 'current_rows') or not self.current_rows:
            QMessageBox.warning(self, "No Data", "There are no stock alerts to export.")
            return
        
        lang_code = lang.get_current()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"stock_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Stock Alerts" if lang_code != "my" else "စတော့သတိပေးချက်စာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            symbol = get_currency_symbol()
            rows = self.current_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Alerts"
            
            # Title
            ws.merge_cells('A1:J1')
            if lang_code == "my":
                title_text = "စတော့သတိပေးချက်အစီရင်ခံစာ"
            else:
                title_text = "STOCK ALERT REPORT"
            ws['A1'] = title_text
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Subtitle
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Alerts: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Filter info
            search_text = self.search_input.text().strip()
            status_filter = self.status_filter.currentText()
            filter_info = []
            if search_text:
                filter_info.append(f"Search: {search_text}")
            if status_filter and status_filter not in ["All Status", "အားလုံး"]:
                filter_info.append(f"Status: {status_filter}")
            if filter_info:
                ws['A4'] = " | ".join(filter_info)
                ws['A4'].font = Font(size=9, italic=True, color="7f8c8d")
                start_row = 6
            else:
                start_row = 5
            
            # Headers
            if lang_code == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", 
                          "လက်ကျန်", "သတိပေးပမာဏ", "အခြေအနေ", "ပြန်မှာသင့်ပမာဏ", 
                          "စျေးနှုန်း", "စတော့တန်ဖိုး"]
            else:
                headers = ["Product Name", "SKU", "Barcode", "Category", 
                          "Current Stock", "Low Stock Level", "Status", "Suggested Order", 
                          "Price", "Stock Value"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            total_stock_value = 0
            total_low_products = 0
            total_out_of_stock = 0
            
            for row_idx, row_data in enumerate(rows, start=start_row + 1):
                product_id, name, sku, stock, low_stock, sold_by, status, suggested, price, category, barcode = row_data
                
                stock_val = int(stock) if stock else 0
                price_val = float(price) if price else 0
                stock_value = price_val * stock_val
                
                ws.cell(row=row_idx, column=1, value=name or "")
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=barcode or "")
                ws.cell(row=row_idx, column=4, value=category or "")
                ws.cell(row=row_idx, column=5, value=stock_val)
                ws.cell(row=row_idx, column=6, value=low_stock or 0)
                ws.cell(row=row_idx, column=7, value=status or "")
                ws.cell(row=row_idx, column=8, value=suggested or 0)
                ws.cell(row=row_idx, column=9, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=10, value=format_money(stock_value, symbol))
                
                total_stock_value += stock_value
                if status == "Out of Stock":
                    total_out_of_stock += 1
                elif status == "Low Stock":
                    total_low_products += 1
                
                # Color code status column
                status_cell = ws.cell(row=row_idx, column=7)
                if status == "Out of Stock":
                    status_cell.fill = PatternFill(start_color="ff4444", end_color="ff4444", fill_type="solid")
                    status_cell.font = Font(color="FFFFFF")
                elif status == "Low Stock":
                    status_cell.fill = PatternFill(start_color="ffcc00", end_color="ffcc00", fill_type="solid")
                
                # Highlight suggested order column
                if suggested > 0:
                    suggested_cell = ws.cell(row=row_idx, column=8)
                    suggested_cell.fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
            
            # Summary section
            summary_row = len(rows) + start_row + 3
            ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
            
            summary_data = [
                ("Total Products with Alerts", len(rows)),
                ("Out of Stock", total_out_of_stock),
                ("Low Stock", total_low_products),
                ("Total Stock Value", format_money(total_stock_value, symbol)),
                ("", ""),
                ("Note: Suggested order = Low Stock Level × 2", "")
            ]
            
            for i, (label, value) in enumerate(summary_data):
                row = summary_row + 2 + i
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                if "Note:" in label or "မှတ်ချက်" in label:
                    ws.cell(row=row, column=1, value=label).font = Font(italic=True, color="7f8c8d")
                else:
                    ws.cell(row=row, column=2, value=value)
            
            # Auto adjust columns
            for col in range(1, 11):
                column_letter = chr(64 + col)
                ws.column_dimensions[column_letter].width = 18
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            logger.error(f"Export stock alerts failed: {e}")
            ExcelExporter.show_error_message(self, e)

    def keyPressEvent(self, event):
        """Handle keyboard shortcuts"""
        if event.key() == Qt.Key.Key_Escape:
            self.accept()
        elif event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.key() == Qt.Key.Key_E:
                self.export_to_excel()
            elif event.key() == Qt.Key.Key_F:
                self.search_input.setFocus()
                self.search_input.selectAll()
        super().keyPressEvent(event)