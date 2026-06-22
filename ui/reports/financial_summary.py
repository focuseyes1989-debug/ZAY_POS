# ui/reports/financial_summary.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget, QGroupBox
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from loguru import logger
from datetime import datetime
import csv


class FinancialSummaryWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Sales data
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0), COUNT(*), COALESCE(AVG(total), 0)
                FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_sales, transaction_count, avg_sale = cursor.fetchone()
            
            # COGS
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (self.from_date, self.to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Expenses
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0), COUNT(*)
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_expenses, expense_count = cursor.fetchone()
            
            # Sales by category
            cursor.execute("""
                SELECT COALESCE(p.category, 'Uncategorized') as category,
                       COALESCE(SUM(si.total), 0) as sales,
                       COALESCE(SUM(p.cost * si.qty), 0) as cogs
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                LEFT JOIN products p ON si.product_name = p.name
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY p.category
                ORDER BY sales DESC
            """, (self.from_date, self.to_date))
            sales_categories = cursor.fetchall()
            
            # Expenses by category
            cursor.execute("""
                SELECT category, COALESCE(SUM(amount), 0) as total
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                GROUP BY category
                ORDER BY total DESC
            """, (self.from_date, self.to_date))
            expense_categories = cursor.fetchall()
            
            conn.close()
            
            gross_profit = (total_sales or 0) - (total_cogs or 0)
            net_profit = gross_profit - (total_expenses or 0)
            net_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            self.result.emit({
                'total_sales': total_sales or 0,
                'total_cogs': total_cogs or 0,
                'gross_profit': gross_profit,
                'total_expenses': total_expenses or 0,
                'net_profit': net_profit,
                'net_margin': net_margin,
                'transaction_count': transaction_count,
                'expense_count': expense_count,
                'sales_categories': sales_categories,
                'expense_categories': expense_categories
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class FinancialSummaryTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Row 1
        card_layout1 = QHBoxLayout()
        card_layout1.setSpacing(15)
        
        self.sales_card = self.parent_dialog.create_card("Total Sales")
        card_layout1.addWidget(self.sales_card, 1)
        
        self.cogs_card = self.parent_dialog.create_card("COGS", color="#e74c3c")
        card_layout1.addWidget(self.cogs_card, 1)
        
        self.gross_card = self.parent_dialog.create_card("Gross Profit", color="#2ecc71")
        card_layout1.addWidget(self.gross_card, 1)
        
        layout.addLayout(card_layout1)
        
        # Row 2
        card_layout2 = QHBoxLayout()
        card_layout2.setSpacing(15)
        
        self.expense_card = self.parent_dialog.create_card("Expenses", color="#e67e22")
        card_layout2.addWidget(self.expense_card, 1)
        
        self.net_card = self.parent_dialog.create_card("Net Profit", color="#9b59b6")
        card_layout2.addWidget(self.net_card, 1)
        
        self.margin_card = self.parent_dialog.create_card("Net Margin", color="#1abc9c")
        card_layout2.addWidget(self.margin_card, 1)
        
        layout.addLayout(card_layout2)
        
        # Category breakdown tables
        split_layout = QHBoxLayout()
        split_layout.setSpacing(15)
        
        # Sales by category
        sales_cat_group = QGroupBox("Sales by Product Category")
        sales_cat_layout = QVBoxLayout()
        self.sales_category_table = QTableWidget()
        self.sales_category_table.setColumnCount(3)
        self.sales_category_table.setHorizontalHeaderLabels(["Category", "Sales", "COGS"])
        self.sales_category_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        sales_cat_layout.addWidget(self.sales_category_table)
        sales_cat_group.setLayout(sales_cat_layout)
        split_layout.addWidget(sales_cat_group, 1)
        
        # Expenses by category
        exp_cat_group = QGroupBox("Expenses by Category")
        exp_cat_layout = QVBoxLayout()
        self.expense_category_table = QTableWidget()
        self.expense_category_table.setColumnCount(2)
        self.expense_category_table.setHorizontalHeaderLabels(["Category", "Amount"])
        self.expense_category_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        exp_cat_layout.addWidget(self.expense_category_table)
        exp_cat_group.setLayout(exp_cat_layout)
        split_layout.addWidget(exp_cat_group, 1)
        
        layout.addLayout(split_layout)
        
        self.setLayout(layout)
    
    def refresh(self, from_date, to_date):
        if self._is_loading:
            return
        
        self._is_loading = True
        self.sales_card.amount_label.setText("Loading...")
        self.cogs_card.amount_label.setText("Loading...")
        self.gross_card.amount_label.setText("Loading...")
        self.expense_card.amount_label.setText("Loading...")
        self.net_card.amount_label.setText("Loading...")
        self.margin_card.amount_label.setText("Loading...")
        self.sales_category_table.setRowCount(0)
        self.expense_category_table.setRowCount(0)
        
        worker = FinancialSummaryWorker(from_date, to_date)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        
        self._current_data = result
        self._is_loading = False
        
        self.parent_dialog.update_card(self.sales_card, result['total_sales'], symbol)
        self.parent_dialog.update_card(self.cogs_card, result['total_cogs'], symbol)
        self.parent_dialog.update_card(self.gross_card, result['gross_profit'], symbol)
        self.parent_dialog.update_card(self.expense_card, result['total_expenses'], symbol)
        self.parent_dialog.update_card(self.net_card, result['net_profit'], symbol)
        self.margin_card.amount_label.setText(f"{result['net_margin']:.1f}%")
        
        # Color coding
        if result['net_profit'] >= 0:
            self.net_card.amount_label.setStyleSheet("color: #2ecc71; font-size: 22pt; font-weight: bold;")
            self.margin_card.amount_label.setStyleSheet("color: #2ecc71; font-size: 22pt; font-weight: bold;")
        else:
            self.net_card.amount_label.setStyleSheet("color: #e74c3c; font-size: 22pt; font-weight: bold;")
            self.margin_card.amount_label.setStyleSheet("color: #e74c3c; font-size: 22pt; font-weight: bold;")
        
        # Sales by category
        self.sales_category_table.setRowCount(len(result['sales_categories']))
        for i, (cat, sales, cogs) in enumerate(result['sales_categories']):
            self.sales_category_table.setItem(i, 0, QTableWidgetItem(cat or "Uncategorized"))
            self.sales_category_table.setItem(i, 1, QTableWidgetItem(format_money(sales, symbol)))
            self.sales_category_table.setItem(i, 2, QTableWidgetItem(format_money(cogs, symbol)))
        
        # Expenses by category
        self.expense_category_table.setRowCount(len(result['expense_categories']))
        for i, (cat, amount) in enumerate(result['expense_categories']):
            self.expense_category_table.setItem(i, 0, QTableWidgetItem(cat))
            self.expense_category_table.setItem(i, 1, QTableWidgetItem(format_money(amount, symbol)))
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"financial_summary_{from_date}_to_{to_date}.xlsx",
            "Export Financial Summary"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Summary data
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (from_date, to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Sales by category
            cursor.execute("""
                SELECT COALESCE(p.category, 'Uncategorized') as category,
                       COALESCE(SUM(si.total), 0) as sales,
                       COALESCE(SUM(p.cost * si.qty), 0) as cogs
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                LEFT JOIN products p ON si.product_name = p.name
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY p.category
                ORDER BY sales DESC
            """, (from_date, to_date))
            sales_categories = cursor.fetchall()
            
            # Expenses by category
            cursor.execute("""
                SELECT category, COALESCE(SUM(amount), 0) as total
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                GROUP BY category
                ORDER BY total DESC
            """, (from_date, to_date))
            expense_categories = cursor.fetchall()
            
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Summary"
            
            # Title
            ws.merge_cells('A1:C1')
            ws['A1'] = "FINANCIAL SUMMARY"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            
            summary_data = [
                ("Total Sales", format_money(total_sales, symbol)),
                ("COGS", format_money(total_cogs, symbol)),
                ("Gross Profit", format_money(gross_profit, symbol)),
                ("Total Expenses", format_money(total_expenses, symbol)),
                ("Net Profit", format_money(net_profit, symbol)),
                ("Net Margin", f"{margin:.1f}%"),
            ]
            
            for i, (label, value) in enumerate(summary_data, start=6):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # Sales by category
            row = 14
            ws.cell(row=row, column=1, value="Sales by Category")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Category")
            ws.cell(row=row, column=2, value="Sales")
            ws.cell(row=row, column=3, value="COGS")
            for col in range(1, 4):
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            
            for cat, sales, cogs in sales_categories:
                ws.cell(row=row, column=1, value=cat or "Uncategorized")
                ws.cell(row=row, column=2, value=float(sales))
                ws.cell(row=row, column=3, value=float(cogs))
                row += 1
            
            # Expenses by category
            row += 2
            ws.cell(row=row, column=1, value="Expenses by Category")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Category")
            ws.cell(row=row, column=2, value="Amount")
            for col in range(1, 3):
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            
            for cat, amount in expense_categories:
                ws.cell(row=row, column=1, value=cat)
                ws.cell(row=row, column=2, value=float(amount))
                row += 1
            
            # Auto adjust columns
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)