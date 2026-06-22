# ui/reports/expense_report.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget, QComboBox, QLabel
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from loguru import logger
from datetime import datetime
import csv


class ExpenseReportWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date, category=None):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
        self.category = category
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            query = """
                SELECT expense_date, expense_no, category, description, amount, payment_method
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """
            params = [self.from_date, self.to_date]
            
            if self.category and self.category != "All Categories":
                query += " AND category = ?"
                params.append(self.category)
            
            query += " ORDER BY expense_date DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            total_expenses = sum(row[4] for row in rows) if rows else 0
            expense_count = len(rows)
            avg_expense = total_expenses / expense_count if expense_count > 0 else 0
            
            conn.close()
            
            self.result.emit({
                'rows': rows,
                'total_expenses': total_expenses,
                'expense_count': expense_count,
                'avg_expense': avg_expense
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class ExpenseReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Summary cards
        card_layout = QHBoxLayout()
        card_layout.setSpacing(15)
        
        self.total_card = self.parent_dialog.create_card("Total Expenses", color="#e74c3c")
        card_layout.addWidget(self.total_card, 1)
        
        self.count_card = self.parent_dialog.create_card("Number of Expenses")
        card_layout.addWidget(self.count_card, 1)
        
        self.avg_card = self.parent_dialog.create_card("Average Expense")
        card_layout.addWidget(self.avg_card, 1)
        
        layout.addLayout(card_layout)
        
        # Category filter
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Category:"))
        self.category_filter = QComboBox()
        self.category_filter.addItem("All Categories")
        self.category_filter.currentTextChanged.connect(self.on_category_changed)
        filter_layout.addWidget(self.category_filter)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Date", "Expense No", "Category", "Description", "Amount", "Payment Method"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)
        
        self.load_categories()
        self.setLayout(layout)
    
    def load_categories(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM expense_categories ORDER BY name")
            rows = cursor.fetchall()
            for row in rows:
                self.category_filter.addItem(row[0])
            conn.close()
        except Exception as e:
            logger.error(f"Failed to load categories: {e}")
    
    def on_category_changed(self):
        if not self._is_loading:
            self.parent_dialog.refresh_current_tab()
    
    def refresh(self, from_date, to_date):
        if self._is_loading:
            return
        
        self._is_loading = True
        category = self.category_filter.currentText()
        
        self.table.setRowCount(0)
        self.total_card.amount_label.setText("Loading...")
        self.count_card.amount_label.setText("Loading...")
        self.avg_card.amount_label.setText("Loading...")
        
        worker = ExpenseReportWorker(from_date, to_date, category)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        
        self._current_data = result
        self._is_loading = False
        
        self.parent_dialog.update_card(self.total_card, result['total_expenses'], symbol)
        self.count_card.amount_label.setText(str(result['expense_count']))
        self.parent_dialog.update_card(self.avg_card, result['avg_expense'], symbol)
        
        self.table.setRowCount(len(result['rows']))
        for i, row in enumerate(result['rows']):
            self.table.setItem(i, 0, QTableWidgetItem(row[0] or ""))
            self.table.setItem(i, 1, QTableWidgetItem(row[1] or ""))
            self.table.setItem(i, 2, QTableWidgetItem(row[2] or ""))
            self.table.setItem(i, 3, QTableWidgetItem(row[3] or ""))
            self.table.setItem(i, 4, QTableWidgetItem(format_money(row[4], symbol)))
            self.table.setItem(i, 5, QTableWidgetItem(row[5] or ""))
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"expense_report_{from_date}_to_{to_date}.xlsx",
            "Export Expense Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT expense_date, expense_no, category, description, amount, payment_method
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                ORDER BY expense_date DESC
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            # Get totals
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0), COUNT(*), COALESCE(AVG(amount), 0)
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses, count, avg_expense = cursor.fetchone()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Expense Report"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "EXPENSE REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = f"Total Expenses: {format_money(total_expenses, symbol)}"
            ws['A7'] = f"Transactions: {count}"
            ws['A8'] = f"Average Expense: {format_money(avg_expense, symbol)}"
            
            # Headers
            headers = ["Date", "Expense No", "Category", "Description", "Amount", "Payment Method"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, row in enumerate(rows, start=11):
                ws.cell(row=row_idx, column=1, value=row[0] or "")
                ws.cell(row=row_idx, column=2, value=row[1] or "")
                ws.cell(row=row_idx, column=3, value=row[2] or "")
                ws.cell(row=row_idx, column=4, value=row[3] or "")
                ws.cell(row=row_idx, column=5, value=float(row[4]) if row[4] else 0)
                ws.cell(row=row_idx, column=6, value=row[5] or "")
            
            # Auto adjust columns
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].auto_size = True
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)