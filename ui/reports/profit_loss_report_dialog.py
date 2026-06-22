# ui/reports/profit_loss_report_dialog.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QFrame
from PyQt6.QtCore import Qt, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import format_money
from ui.reports.base_report_dialog import BaseReportDialog
from datetime import datetime
import csv


class ProfitLossWorker(QObject):
    """Worker for calculating profit/loss in background thread"""
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        """Run the calculation in background thread"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total Sales
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_sales = cursor.fetchone()[0]
            
            # COGS
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (self.from_date, self.to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Gross Profit
            gross_profit = total_sales - total_cogs
            
            # Total Expenses
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_expenses = cursor.fetchone()[0]
            
            # Net Profit
            net_profit = gross_profit - total_expenses
            
            # Profit Margin
            profit_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            conn.close()
            
            result = {
                'total_sales': total_sales,
                'total_cogs': total_cogs,
                'gross_profit': gross_profit,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'profit_margin': profit_margin,
            }
            
            self.result.emit(result)
            
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class ProfitLossReportDialog(BaseReportDialog):
    def __init__(self, parent=None):
        super().__init__("Profit & Loss Report", parent)
        self.create_content_area()
        
        # Initial refresh
        self.refresh_report()
        self.retranslateUi()
    
    def create_content_area(self):
        """Create main content area with cards and table"""
        # Summary cards layout
        card_layout = QHBoxLayout()
        card_layout.setSpacing(15)
        
        # Sales Card
        self.sales_card = self.create_card("Total Sales", "0", "#3498db")
        card_layout.addWidget(self.sales_card, 1)
        
        # COGS Card
        self.cogs_card = self.create_card("COGS", "0", "#e74c3c")
        card_layout.addWidget(self.cogs_card, 1)
        
        # Gross Profit Card
        self.gross_card = self.create_card("Gross Profit", "0", "#2ecc71")
        card_layout.addWidget(self.gross_card, 1)
        
        # Expenses Card
        self.expenses_card = self.create_card("Operating Expenses", "0", "#e67e22")
        card_layout.addWidget(self.expenses_card, 1)
        
        # Net Profit Card
        self.net_card = self.create_card("Net Profit", "0", "#9b59b6")
        card_layout.addWidget(self.net_card, 1)
        
        # Margin Card
        self.margin_card = self.create_card("Net Margin", "0%", "#1abc9c")
        card_layout.addWidget(self.margin_card, 1)
        
        self.main_layout.insertLayout(2, card_layout)
        
        # Summary table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Metric", "Amount", "% of Sales", "Status", "Trend"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        
        self.main_layout.insertWidget(3, self.table)
    
    def create_card(self, title, amount, color):
        """Create a summary card"""
        card = QFrame()
        card.setObjectName("reportCard")
        card.setFrameStyle(QFrame.Shape.StyledPanel)
        card.setMinimumHeight(100)
        
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(5)
        
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        amount_label = QLabel(amount)
        amount_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        amount_label.setStyleSheet(f"color: {color}; font-size: 20pt; font-weight: bold;")
        
        layout.addWidget(title_label)
        layout.addWidget(amount_label)
        
        # Store references for updates
        card.amount_label = amount_label
        return card
    
    def update_card(self, card, amount, symbol=None):
        """Update card amount"""
        if symbol:
            card.amount_label.setText(format_money(amount, symbol))
        else:
            card.amount_label.setText(str(amount))
    
    def create_worker(self):
        """Create worker for background refresh"""
        from_date, to_date = self.get_date_range()
        return ProfitLossWorker(from_date, to_date)
    
    def update_ui_with_result(self, result):
        """Update UI with calculation result"""
        symbol = self.get_currency_symbol()
        
        total_sales = result['total_sales']
        total_cogs = result['total_cogs']
        gross_profit = result['gross_profit']
        total_expenses = result['total_expenses']
        net_profit = result['net_profit']
        profit_margin = result['profit_margin']
        
        # Update cards
        self.update_card(self.sales_card, total_sales, symbol)
        self.update_card(self.cogs_card, total_cogs, symbol)
        self.update_card(self.gross_card, gross_profit, symbol)
        self.update_card(self.expenses_card, total_expenses, symbol)
        self.update_card(self.net_card, net_profit, symbol)
        self.update_card(self.margin_card, f"{profit_margin:.1f}%")
        
        # Color coding for net profit
        if net_profit >= 0:
            self.net_card.amount_label.setStyleSheet("color: #2ecc71; font-size: 20pt; font-weight: bold;")
            self.margin_card.amount_label.setStyleSheet("color: #2ecc71; font-size: 20pt; font-weight: bold;")
        else:
            self.net_card.amount_label.setStyleSheet("color: #e74c3c; font-size: 20pt; font-weight: bold;")
            self.margin_card.amount_label.setStyleSheet("color: #e74c3c; font-size: 20pt; font-weight: bold;")
        
        # Update table
        data = [
            ("Total Sales", total_sales, 100, "Income", "↑"),
            ("COGS", total_cogs, (total_cogs/total_sales*100) if total_sales > 0 else 0, "Expense", "↓"),
            ("Gross Profit", gross_profit, (gross_profit/total_sales*100) if total_sales > 0 else 0, 
             "Profit" if gross_profit >= 0 else "Loss", "↑" if gross_profit >= 0 else "↓"),
            ("Operating Expenses", total_expenses, (total_expenses/total_sales*100) if total_sales > 0 else 0, "Expense", "↓"),
            ("Net Profit", net_profit, (net_profit/total_sales*100) if total_sales > 0 else 0,
             "Profit" if net_profit >= 0 else "Loss", "↑" if net_profit >= 0 else "↓"),
        ]
        
        self.table.setRowCount(len(data))
        for i, (metric, amount, percentage, status, trend) in enumerate(data):
            self.table.setItem(i, 0, QTableWidgetItem(metric))
            self.table.setItem(i, 1, QTableWidgetItem(format_money(amount, symbol)))
            self.table.setItem(i, 2, QTableWidgetItem(f"{percentage:.1f}%"))
            
            status_item = QTableWidgetItem(status)
            if status == "Profit":
                status_item.setForeground(QColor(46, 204, 113))
            elif status == "Loss":
                status_item.setForeground(QColor(231, 76, 60))
            self.table.setItem(i, 3, status_item)
            
            trend_item = QTableWidgetItem(trend)
            if trend == "↑":
                trend_item.setForeground(QColor(46, 204, 113))
            elif trend == "↓":
                trend_item.setForeground(QColor(231, 76, 60))
            self.table.setItem(i, 4, trend_item)
    
    def export_to_excel(self):
        """Export to Excel"""
        from_date, to_date = self.get_date_range()
        
        from utils.excel_exporter import ExcelExporter
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"profit_loss_{from_date}_to_{to_date}.xlsx",
            "Export Profit & Loss Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = self.get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_cogs = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses = cursor.fetchone()[0]
            
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            profit_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "PROFIT & LOSS REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Headers
            headers = ["Metric", "Amount", "% of Sales"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            data = [
                ("Total Sales", total_sales, "100%"),
                ("COGS", total_cogs, f"{(total_cogs/total_sales*100):.1f}%" if total_sales > 0 else "0%"),
                ("Gross Profit", gross_profit, f"{(gross_profit/total_sales*100):.1f}%" if total_sales > 0 else "0%"),
                ("Operating Expenses", total_expenses, f"{(total_expenses/total_sales*100):.1f}%" if total_sales > 0 else "0%"),
                ("Net Profit", net_profit, f"{profit_margin:.1f}%"),
            ]
            
            for row_idx, (metric, amount, percentage) in enumerate(data, start=6):
                ws.cell(row=row_idx, column=1, value=metric)
                ws.cell(row=row_idx, column=2, value=format_money(amount, symbol))
                ws.cell(row=row_idx, column=3, value=percentage)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def retranslateUi(self):
        """Update UI text based on language"""
        lang = self.get_lang()
        if lang == "my":
            self.setWindowTitle("အမြတ်အစွန်း အစီရင်ခံစာ")
            self.btn_refresh.setText("ပြန်လည်")
            self.btn_export_excel.setText("📊 Excel ထုတ်မည်")
            self.btn_export_pdf.setText("📄 PDF ထုတ်မည်")
            self.btn_close.setText("ပိတ်မည်")
            self.table.setHorizontalHeaderLabels([
                "အမျိုးအစား", "ပမာဏ", "ရောင်းအား၏ ရာခိုင်နှုန်း", "အခြေအနေ", "လမ်းကြောင်း"
            ])