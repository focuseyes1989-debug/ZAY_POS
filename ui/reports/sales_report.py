# ui/reports/sales_report.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from loguru import logger
import csv


class SalesReportWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT s.created_at, s.invoice_no, c.name, s.total, s.payment, s.change_amount
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                ORDER BY s.created_at DESC
            """, (self.from_date, self.to_date))
            rows = cursor.fetchall()
            
            total_sales = sum(row[3] for row in rows) if rows else 0
            transaction_count = len(rows)
            avg_sale = total_sales / transaction_count if transaction_count > 0 else 0
            
            conn.close()
            
            self.result.emit({
                'rows': rows,
                'total_sales': total_sales,
                'transaction_count': transaction_count,
                'avg_sale': avg_sale
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class SalesReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Summary cards
        card_layout = QHBoxLayout()
        card_layout.setSpacing(15)
        
        self.total_card = self.parent_dialog.create_card("Total Sales")
        card_layout.addWidget(self.total_card, 1)
        
        self.count_card = self.parent_dialog.create_card("Transactions")
        card_layout.addWidget(self.count_card, 1)
        
        self.avg_card = self.parent_dialog.create_card("Average Sale")
        card_layout.addWidget(self.avg_card, 1)
        
        layout.addLayout(card_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Date", "Invoice No", "Customer", "Total", "Payment", "Change"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def refresh(self, from_date, to_date):
        if self._is_loading:
            return
        
        self._is_loading = True
        self.table.setRowCount(0)
        self.total_card.amount_label.setText("Loading...")
        self.count_card.amount_label.setText("Loading...")
        self.avg_card.amount_label.setText("Loading...")
        
        worker = SalesReportWorker(from_date, to_date)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        
        self._current_data = result
        self._is_loading = False
        
        self.parent_dialog.update_card(self.total_card, result['total_sales'], symbol)
        self.count_card.amount_label.setText(str(result['transaction_count']))
        self.parent_dialog.update_card(self.avg_card, result['avg_sale'], symbol)
        
        self.table.setRowCount(len(result['rows']))
        for i, row in enumerate(result['rows']):
            self.table.setItem(i, 0, QTableWidgetItem(str(row[0])[:16] if row[0] else ""))
            self.table.setItem(i, 1, QTableWidgetItem(row[1] or ""))
            self.table.setItem(i, 2, QTableWidgetItem(row[2] if row[2] else "Walk-in"))
            self.table.setItem(i, 3, QTableWidgetItem(format_money(row[3], symbol)))
            self.table.setItem(i, 4, QTableWidgetItem(format_money(row[4], symbol)))
            self.table.setItem(i, 5, QTableWidgetItem(format_money(row[5], symbol)))
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"sales_report_{from_date}_to_{to_date}.xlsx",
            "Export Sales Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT s.created_at, s.invoice_no, c.name, s.total, s.payment, s.change_amount
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                ORDER BY s.created_at DESC
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            # Get totals
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0), COUNT(*), COALESCE(AVG(total), 0)
                FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales, count, avg_sale = cursor.fetchone()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "SALES REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = f"Total Sales: {format_money(total_sales, symbol)}"
            ws['A7'] = f"Transactions: {count}"
            ws['A8'] = f"Average Sale: {format_money(avg_sale, symbol)}"
            
            # Headers
            headers = ["Date", "Invoice No", "Customer", "Total", "Payment", "Change"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, row in enumerate(rows, start=11):
                ws.cell(row=row_idx, column=1, value=str(row[0]) if row[0] else "")
                ws.cell(row=row_idx, column=2, value=row[1] or "")
                ws.cell(row=row_idx, column=3, value=row[2] if row[2] else "Walk-in")
                ws.cell(row=row_idx, column=4, value=float(row[3]) if row[3] else 0)
                ws.cell(row=row_idx, column=5, value=float(row[4]) if row[4] else 0)
                ws.cell(row=row_idx, column=6, value=float(row[5]) if row[5] else 0)
            
            # Auto adjust columns
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].auto_size = True
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)