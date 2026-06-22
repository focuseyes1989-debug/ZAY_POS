# ui/reports/profit_loss_report.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from loguru import logger
from datetime import datetime
import csv


class PLReportWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total sales
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_sales = cursor.fetchone()[0]
            
            # COGS
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (self.from_date, self.to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Expenses
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_expenses = cursor.fetchone()[0]
            
            # Monthly breakdown
            cursor.execute("""
                SELECT 
                    strftime('%Y-%m', s.created_at) as month,
                    COALESCE(SUM(s.total), 0) as sales,
                    COALESCE(SUM(p.cost * si.qty), 0) as cogs,
                    COALESCE(SUM(e.amount), 0) as expenses
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                LEFT JOIN products p ON si.product_name = p.name
                LEFT JOIN expenses e ON strftime('%Y-%m', e.expense_date) = strftime('%Y-%m', s.created_at)
                WHERE s.status = 'completed' 
                  AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY strftime('%Y-%m', s.created_at)
                ORDER BY month
            """, (self.from_date, self.to_date))
            monthly_rows = cursor.fetchall()
            
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            net_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            self.result.emit({
                'total_sales': total_sales,
                'total_cogs': total_cogs,
                'gross_profit': gross_profit,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'net_margin': net_margin,
                'monthly_rows': monthly_rows
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class ProfitLossReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        # Row 1
        card_layout1 = QHBoxLayout()
        card_layout1.setSpacing(15)
        
        self.sales_card = self.parent_dialog.create_card("Total Sales (Income)")
        card_layout1.addWidget(self.sales_card, 1)
        
        self.cogs_card = self.parent_dialog.create_card("COGS", color="#e74c3c")
        card_layout1.addWidget(self.cogs_card, 1)
        
        self.gross_card = self.parent_dialog.create_card("Gross Profit", color="#2ecc71")
        card_layout1.addWidget(self.gross_card, 1)
        
        layout.addLayout(card_layout1)
        
        # Row 2
        card_layout2 = QHBoxLayout()
        card_layout2.setSpacing(15)
        
        self.expense_card = self.parent_dialog.create_card("Operating Expenses", color="#e67e22")
        card_layout2.addWidget(self.expense_card, 1)
        
        self.net_card = self.parent_dialog.create_card("Net Profit", color="#9b59b6")
        card_layout2.addWidget(self.net_card, 1)
        
        self.margin_card = self.parent_dialog.create_card("Net Margin", color="#1abc9c")
        card_layout2.addWidget(self.margin_card, 1)
        
        layout.addLayout(card_layout2)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Month", "Sales", "COGS", "Gross Profit", "Expenses", "Net Profit", "Margin %"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def refresh(self, from_date, to_date):
        if self._is_loading:
            return
        
        self._is_loading = True
        self.table.setRowCount(0)
        self.sales_card.amount_label.setText("Loading...")
        self.cogs_card.amount_label.setText("Loading...")
        self.gross_card.amount_label.setText("Loading...")
        self.expense_card.amount_label.setText("Loading...")
        self.net_card.amount_label.setText("Loading...")
        self.margin_card.amount_label.setText("Loading...")
        
        worker = PLReportWorker(from_date, to_date)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        
        self._current_data = result
        self._is_loading = False
        
        self.parent_dialog.update_card(self.sales_card, result['total_sales'], symbol)
        self.parent_dialog.update_card(self.cogs_card, result['total_cogs'], symbol)
        self.parent_dialog.update_card(self.gross_card, result['gross_profit'], symbol)
        self.parent_dialog.update_card(self.expense_card, result['total_expenses'], symbol)
        self.parent_dialog.update_card(self.net_card, result['net_profit'], symbol)
        self.margin_card.amount_label.setText(f"{result['net_margin']:.1f}%")
        
        # Color coding
        if result['net_profit'] >= 0:
            self.net_card.amount_label.setStyleSheet("color: #2ecc71; font-size: 22pt; font-weight: bold;")
            self.margin_card.amount_label.setStyleSheet("color: #2ecc71; font-size: 22pt; font-weight: bold;")
        else:
            self.net_card.amount_label.setStyleSheet("color: #e74c3c; font-size: 22pt; font-weight: bold;")
            self.margin_card.amount_label.setStyleSheet("color: #e74c3c; font-size: 22pt; font-weight: bold;")
        
        # Monthly breakdown
        rows = result['monthly_rows']
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            month, sales, cogs, expenses = row
            gross = sales - cogs
            net = gross - expenses
            margin = (net / sales * 100) if sales > 0 else 0
            
            self.table.setItem(i, 0, QTableWidgetItem(month))
            self.table.setItem(i, 1, QTableWidgetItem(format_money(sales, symbol)))
            self.table.setItem(i, 2, QTableWidgetItem(format_money(cogs, symbol)))
            
            gross_item = QTableWidgetItem(format_money(gross, symbol))
            gross_item.setForeground(QColor(46, 204, 113) if gross >= 0 else QColor(231, 76, 60))
            self.table.setItem(i, 3, gross_item)
            
            self.table.setItem(i, 4, QTableWidgetItem(format_money(expenses, symbol)))
            
            net_item = QTableWidgetItem(format_money(net, symbol))
            net_item.setForeground(QColor(46, 204, 113) if net >= 0 else QColor(231, 76, 60))
            self.table.setItem(i, 5, net_item)
            
            margin_item = QTableWidgetItem(f"{margin:.1f}%")
            if margin >= 10:
                margin_item.setForeground(QColor(46, 204, 113))
            elif margin >= 0:
                margin_item.setForeground(QColor(241, 196, 15))
            else:
                margin_item.setForeground(QColor(231, 76, 60))
            self.table.setItem(i, 6, margin_item)
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"profit_loss_report_{from_date}_to_{to_date}.xlsx",
            "Export Profit & Loss Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    strftime('%Y-%m', s.created_at) as month,
                    COALESCE(SUM(s.total), 0) as sales,
                    COALESCE(SUM(p.cost * si.qty), 0) as cogs,
                    COALESCE(SUM(e.amount), 0) as expenses
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                LEFT JOIN products p ON si.product_name = p.name
                LEFT JOIN expenses e ON strftime('%Y-%m', e.expense_date) = strftime('%Y-%m', s.created_at)
                WHERE s.status = 'completed' 
                  AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY strftime('%Y-%m', s.created_at)
                ORDER BY month
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            # Get totals
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_cogs = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses = cursor.fetchone()[0]
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = "PROFIT & LOSS REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = f"Total Sales: {format_money(total_sales, symbol)}"
            ws['A7'] = f"COGS: {format_money(total_cogs, symbol)}"
            ws['A8'] = f"Gross Profit: {format_money(gross_profit, symbol)}"
            ws['A9'] = f"Total Expenses: {format_money(total_expenses, symbol)}"
            ws['A10'] = f"Net Profit: {format_money(net_profit, symbol)}"
            ws['A11'] = f"Net Margin: {margin:.1f}%"
            
            # Headers
            headers = ["Month", "Sales", "COGS", "Gross Profit", "Expenses", "Net Profit", "Margin %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=13, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, row in enumerate(rows, start=14):
                month, sales, cogs, expenses = row
                gross = sales - cogs
                net = gross - expenses
                margin_pct = (net / sales * 100) if sales > 0 else 0
                
                ws.cell(row=row_idx, column=1, value=month)
                ws.cell(row=row_idx, column=2, value=float(sales))
                ws.cell(row=row_idx, column=3, value=float(cogs))
                ws.cell(row=row_idx, column=4, value=float(gross))
                ws.cell(row=row_idx, column=5, value=float(expenses))
                ws.cell(row=row_idx, column=6, value=float(net))
                ws.cell(row=row_idx, column=7, value=float(margin_pct))
            
            # Auto adjust columns
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].auto_size = True
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)